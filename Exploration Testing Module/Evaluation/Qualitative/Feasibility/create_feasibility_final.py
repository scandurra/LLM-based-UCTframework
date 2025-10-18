#!/usr/bin/env python3
"""
Script finale per creare UC-Analysis-Feasibility.xlsx con design ottimizzato
"""

import json
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError as e:
    print(f"Errore importazione: {e}")
    print("Installare con: pip3 install openpyxl")
    sys.exit(1)

import glob

def read_json_file(file_path):
    """Legge un file JSON e restituisce i test case"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if isinstance(data, list):
                return data
            else:
                return [data] if data else []
    except Exception as e:
        print(f"Errore nella lettura di {file_path}: {e}")
        return []

def get_organized_data():
    """Raccoglie e organizza tutti i dati dalle generazioni"""
    base_path = "/Users/raphaelmazzoleni/Desktop/RaphaelMac/ExplorationModule/Con Evaluation/Generations"
    
    generations = [
        "R1 Zero Shot", "R1 One Shot", "R1 Few Shot",
        "R2 Zero Shot", "R2 One Shot", "R2 Few Shot", 
        "R3 Zero Shot", "R3 One Shot", "R3 Few Shot"
    ]
    
    organized_data = []
    all_test_cases = set()
    
    print("Raccolta e organizzazione dati...")
    
    for generation in generations:
        gen_path = os.path.join(base_path, generation)
        if not os.path.exists(gen_path):
            continue
            
        print(f"Processando {generation}...")
        gen_data = {"generation": generation, "use_cases": []}
        
        # Trova tutti i file JSON
        json_files = glob.glob(os.path.join(gen_path, "*.json"))
        json_files.sort()  # Ordina i file
        
        for json_file in json_files:
            uc_name = os.path.basename(json_file).replace('.json', '')
            test_cases = read_json_file(json_file)
            
            if test_cases:  # Solo se ci sono test cases
                uc_test_cases = []
                for tc in test_cases:
                    tc_id = tc.get('test_case_id', '')
                    if tc_id:
                        uc_test_cases.append(tc_id)
                        all_test_cases.add(tc_id)
                
                if uc_test_cases:
                    gen_data["use_cases"].append({
                        "use_case_id": uc_name,
                        "test_cases": uc_test_cases
                    })
        
        if gen_data["use_cases"]:
            organized_data.append(gen_data)
    
    return organized_data, sorted(all_test_cases)

def create_enhanced_excel():
    """Crea il file Excel con design ottimizzato"""
    
    organized_data, all_test_cases = get_organized_data()
    
    print(f"Totale test case unici: {len(all_test_cases)}")
    
    # Crea workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "UC Feasibility Analysis"
    
    # Stili avanzati
    # Headers principali
    main_header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    main_header_font = Font(color="FFFFFF", bold=True, size=11)
    
    # Test case headers
    tc_header_fill = PatternFill(start_color="2F5F8F", end_color="2F5F8F", fill_type="solid")
    tc_header_font = Font(color="FFFFFF", bold=True, size=8)
    
    # Generazioni
    generation_colors = {
        "R1 Zero Shot": "E8F4FD", "R1 One Shot": "E1F0FA", "R1 Few Shot": "D9ECF7",
        "R2 Zero Shot": "FFF2CC", "R2 One Shot": "FFE699", "R2 Few Shot": "FFD966",
        "R3 Zero Shot": "F4CCCC", "R3 One Shot": "EA9999", "R3 Few Shot": "E06666"
    }
    
    # Use Case
    use_case_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    use_case_font = Font(bold=True, size=9)
    
    # Bordi
    thick_border = Border(
        left=Side(style='thick', color='000000'),
        right=Side(style='thick', color='000000'),
        top=Side(style='thick', color='000000'),
        bottom=Side(style='thick', color='000000')
    )
    
    medium_border = Border(
        left=Side(style='medium', color='666666'),
        right=Side(style='medium', color='666666'),
        top=Side(style='medium', color='666666'),
        bottom=Side(style='medium', color='666666')
    )
    
    # Headers
    ws.cell(row=1, column=1, value="Generation")
    ws.cell(row=1, column=2, value="Use Case ID")
    
    # Styling headers principali
    for col in [1, 2]:
        cell = ws.cell(row=1, column=col)
        cell.fill = main_header_fill
        cell.font = main_header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thick_border
    
    # Headers test cases
    col_start = 3
    for i, tc_id in enumerate(all_test_cases):
        col = col_start + i
        cell = ws.cell(row=1, column=col, value=tc_id)
        cell.fill = tc_header_fill
        cell.font = tc_header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
        cell.border = thick_border
    
    current_row = 2
    
    # Popola i dati
    for gen_data in organized_data:
        generation = gen_data["generation"]
        use_cases = gen_data["use_cases"]
        
        if not use_cases:
            continue
            
        generation_start_row = current_row
        
        # Per ogni use case
        for i, uc_data in enumerate(use_cases):
            use_case_id = uc_data["use_case_id"]
            uc_test_cases = uc_data["test_cases"]
            
            # Generation cell (merge per tutte le righe del generation)
            if i == 0:
                ws.cell(row=current_row, column=1, value=generation)
                if len(use_cases) > 1:
                    ws.merge_cells(start_row=current_row, start_column=1,
                                  end_row=current_row + len(use_cases) - 1, end_column=1)
                
                gen_cell = ws.cell(row=current_row, column=1)
                gen_cell.fill = PatternFill(
                    start_color=generation_colors.get(generation, "F0F0F0"),
                    end_color=generation_colors.get(generation, "F0F0F0"),
                    fill_type="solid"
                )
                gen_cell.font = Font(bold=True, size=10)
                gen_cell.alignment = Alignment(horizontal='center', vertical='center')
                gen_cell.border = thick_border
            
            # Use Case cell
            uc_cell = ws.cell(row=current_row, column=2, value=use_case_id)
            uc_cell.fill = use_case_fill
            uc_cell.font = use_case_font
            uc_cell.alignment = Alignment(horizontal='center', vertical='center')
            uc_cell.border = thick_border
            
            # Test case cells
            for j, tc_id in enumerate(all_test_cases):
                col = col_start + j
                tc_cell = ws.cell(row=current_row, column=col, value="")
                
                # Se questo test case appartiene a questo use case
                if tc_id in uc_test_cases:
                    # Colore più evidente per i test case presenti
                    tc_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    tc_cell.border = thick_border
                else:
                    # Bordo meno evidente per i test case non presenti
                    tc_cell.border = medium_border
                
                tc_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Dropdown Yes/No per tutti i test case
                try:
                    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
                    ws.add_data_validation(dv)
                    dv.add(tc_cell)
                except:
                    pass
            
            current_row += 1
    
    # Ottimizzazione colonne
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    
    # Test case columns
    for col in range(col_start, col_start + len(all_test_cases)):
        column_letter = ws.cell(row=1, column=col).column_letter
        ws.column_dimensions[column_letter].width = 3
    
    # Freeze panes
    ws.freeze_panes = 'C2'
    
    # Salva
    output_path = "/Users/raphaelmazzoleni/Desktop/RaphaelMac/ExplorationModule/Con Evaluation/Evaluation/Qualitative/Feasibility/UC-Analysis-Feasibility.xlsx"
    wb.save(output_path)
    
    print(f"File Excel creato: {output_path}")
    print(f"Generazioni processate: {len(organized_data)}")
    print(f"Test case totali: {len(all_test_cases)}")
    print(f"Righe dati: {current_row - 2}")

if __name__ == "__main__":
    create_enhanced_excel()

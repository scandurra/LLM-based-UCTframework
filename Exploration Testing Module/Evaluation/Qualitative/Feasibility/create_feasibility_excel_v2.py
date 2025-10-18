#!/usr/bin/env python3
"""
Script migliorato per creare il file UC-Analysis-Feasibility.xlsx con tutti i test case delle generazioni
"""

import json
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError as e:
    print(f"Errore importazione: {e}")
    print("Installare con: pip3 install openpyxl")
    sys.exit(1)

import glob

def read_json_file(file_path):
    """Legge un file JSON e restituisce i test case"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if isinstance(data, list):
                return data
            else:
                return [data] if data else []
    except Exception as e:
        print(f"Errore nella lettura di {file_path}: {e}")
        return []

def get_all_generations_data():
    """Raccoglie tutti i dati da tutte le generazioni"""
    base_path = "/Users/raphaelmazzoleni/Desktop/RaphaelMac/ExplorationModule/Con Evaluation/Generations"
    
    generations = [
        "R1 Zero Shot", "R1 One Shot", "R1 Few Shot",
        "R2 Zero Shot", "R2 One Shot", "R2 Few Shot", 
        "R3 Zero Shot", "R3 One Shot", "R3 Few Shot"
    ]
    
    data = {}
    all_use_cases = set()
    test_case_headers = []
    
    print("Raccolta dati dalle generazioni...")
    
    for generation in generations:
        gen_path = os.path.join(base_path, generation)
        if not os.path.exists(gen_path):
            print(f"Generazione {generation} non trovata, salto...")
            continue
            
        print(f"Processando {generation}...")
        data[generation] = {}
        
        # Trova tutti i file JSON nella cartella della generazione
        json_files = glob.glob(os.path.join(gen_path, "*.json"))
        
        for json_file in json_files:
            uc_name = os.path.basename(json_file).replace('.json', '')
            test_cases = read_json_file(json_file)
            
            data[generation][uc_name] = []
            all_use_cases.add(uc_name)
            
            for tc in test_cases:
                tc_id = tc.get('test_case_id', '')
                if tc_id:
                    data[generation][uc_name].append(tc_id)
                    if tc_id not in test_case_headers:
                        test_case_headers.append(tc_id)
    
    return data, sorted(all_use_cases), test_case_headers, generations

def create_excel_file():
    """Crea il file Excel con la struttura richiesta"""
    
    # Raccoglie tutti i dati
    data, all_use_cases, test_case_headers, generations = get_all_generations_data()
    
    print(f"Trovati {len(all_use_cases)} use case")
    print(f"Trovati {len(test_case_headers)} test case totali")
    
    # Crea un nuovo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Feasibility Analysis"
    
    # Definisce gli stili
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    
    generation_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    generation_font = Font(bold=True, size=10)
    
    use_case_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    use_case_font = Font(bold=True, size=9)
    
    thick_border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )
    
    # Headers principali
    ws.cell(row=1, column=1, value="Generation")
    ws.cell(row=1, column=2, value="Use Case ID")
    
    # Applica stile agli headers principali
    for col in range(1, 3):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thick_border
    
    # Headers dei test case
    col_offset = 3
    for i, tc_id in enumerate(test_case_headers):
        col = col_offset + i
        header_cell = ws.cell(row=1, column=col, value=tc_id)
        header_cell.fill = header_fill
        header_cell.font = Font(color="FFFFFF", bold=True, size=8)
        header_cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
        header_cell.border = thick_border
    
    current_row = 2
    
    print("Creazione file Excel...")
    
    # Per ogni generazione
    for generation in generations:
        if generation not in data:
            continue
            
        # Raccoglie tutti i use case per questa generazione che hanno test cases
        use_cases_in_gen = sorted([uc for uc in data[generation].keys() if data[generation][uc]])
        
        if not use_cases_in_gen:
            continue
        
        generation_start_row = current_row
        
        # Per ogni use case in questa generazione
        for i, use_case in enumerate(use_cases_in_gen):
            test_cases = data[generation][use_case]
            
            if i == 0:
                # Prima riga della generazione - scrive il nome della generazione
                ws.cell(row=current_row, column=1, value=generation)
                
                # Merge cells per la generazione
                if len(use_cases_in_gen) > 1:
                    ws.merge_cells(start_row=current_row, start_column=1, 
                                  end_row=current_row + len(use_cases_in_gen) - 1, end_column=1)
                
                gen_cell = ws.cell(row=current_row, column=1)
                gen_cell.fill = generation_fill
                gen_cell.font = generation_font
                gen_cell.alignment = Alignment(horizontal='center', vertical='center')
                gen_cell.border = thick_border
            
            # Use Case ID
            ws.cell(row=current_row, column=2, value=use_case)
            uc_cell = ws.cell(row=current_row, column=2)
            uc_cell.fill = use_case_fill
            uc_cell.font = use_case_font
            uc_cell.alignment = Alignment(horizontal='center', vertical='center')
            uc_cell.border = thick_border
            
            # Celle per i test cases
            for j, tc_id in enumerate(test_case_headers):
                col = col_offset + j
                tc_cell = ws.cell(row=current_row, column=col, value="")
                
                # Verifica se questo test case appartiene a questo use case in questa generazione
                if tc_id in test_cases:
                    # Test case presente - colore di sfondo diverso per evidenziare
                    tc_cell.fill = PatternFill(start_color="E7F3E7", end_color="E7F3E7", fill_type="solid")
                
                tc_cell.border = thick_border
                tc_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Validazione dati per Yes/No solo per le celle con test case
                try:
                    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
                    ws.add_data_validation(dv)
                    dv.add(tc_cell)
                except Exception as e:
                    pass
            
            current_row += 1
    
    # Aggiusta la larghezza delle colonne
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    
    # Colonne dei test case più strette
    for col in range(3, len(test_case_headers) + 3):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 2.5
    
    # Congela le prime due colonne e la prima riga
    ws.freeze_panes = 'C2'
    
    # Salva il file
    output_path = "/Users/raphaelmazzoleni/Desktop/RaphaelMac/ExplorationModule/Con Evaluation/Evaluation/Qualitative/Feasibility/UC-Analysis-Feasibility.xlsx"
    wb.save(output_path)
    print(f"File Excel creato: {output_path}")
    print(f"Righe totali: {current_row - 1}")
    print(f"Colonne totali: {len(test_case_headers) + 2}")
    
    return output_path

if __name__ == "__main__":
    create_excel_file()

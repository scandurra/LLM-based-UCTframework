#!/usr/bin/env python3
"""
Script per creare UC-Analysis-Feasibility.xlsx con TC1-TC10 e design corretto
"""

import json
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError as e:
    print(f"Errore importazione: {e}")
    print("Installare con: pip3 install openpyxl")
    sys.exit(1)

import glob

def read_json_file(file_path):
    """Legge un file JSON e restituisce i test case"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if isinstance(data, list):
                return data
            else:
                return [data] if data else []
    except Exception as e:
        print(f"Errore nella lettura di {file_path}: {e}")
        return []

def get_organized_data():
    """Raccoglie e organizza tutti i dati dalle generazioni"""
    base_path = "/Users/raphaelmazzoleni/Desktop/RaphaelMac/ExplorationModule/Con Evaluation/Generations"
    
    generations = [
        "R1 Zero Shot", "R1 One Shot", "R1 Few Shot",
        "R2 Zero Shot", "R2 One Shot", "R2 Few Shot", 
        "R3 Zero Shot", "R3 One Shot", "R3 Few Shot"
    ]
    
    organized_data = []
    
    print("Raccolta e organizzazione dati...")
    
    for generation in generations:
        gen_path = os.path.join(base_path, generation)
        if not os.path.exists(gen_path):
            continue
            
        print(f"Processando {generation}...")
        gen_data = {"generation": generation, "use_cases": []}
        
        # Trova tutti i file JSON
        json_files = glob.glob(os.path.join(gen_path, "*.json"))
        json_files.sort()  # Ordina i file
        
        for json_file in json_files:
            uc_name = os.path.basename(json_file).replace('.json', '')
            test_cases = read_json_file(json_file)
            
            if test_cases:  # Solo se ci sono test cases
                # Conta quanti test case ha questo use case (max 10)
                num_test_cases = min(len(test_cases), 10)
                
                gen_data["use_cases"].append({
                    "use_case_id": uc_name,
                    "num_test_cases": num_test_cases
                })
        
        if gen_data["use_cases"]:
            organized_data.append(gen_data)
    
    return organized_data

def create_feasibility_excel():
    """Crea il file Excel con design ottimizzato e solo TC1-TC10"""
    
    organized_data = get_organized_data()
    
    # Crea workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "UC Feasibility Analysis"
    
    # Stili
    # Headers principali
    main_header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    main_header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # Test case headers
    tc_header_fill = PatternFill(start_color="2F5F8F", end_color="2F5F8F", fill_type="solid")
    tc_header_font = Font(color="FFFFFF", bold=True, size=10)
    
    # Generazioni
    generation_colors = {
        "R1 Zero Shot": "E8F4FD", "R1 One Shot": "E1F0FA", "R1 Few Shot": "D9ECF7",
        "R2 Zero Shot": "FFF2CC", "R2 One Shot": "FFE699", "R2 Few Shot": "FFD966",
        "R3 Zero Shot": "F4CCCC", "R3 One Shot": "EA9999", "R3 Few Shot": "E06666"
    }
    
    # Use Case
    use_case_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    use_case_font = Font(bold=True, size=11)
    
    # Test case presente
    present_tc_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # Bordi
    thick_border = Border(
        left=Side(style='thick', color='000000'),
        right=Side(style='thick', color='000000'),
        top=Side(style='thick', color='000000'),
        bottom=Side(style='thick', color='000000')
    )
    
    thin_border = Border(
        left=Side(style='thin', color='666666'),
        right=Side(style='thin', color='666666'),
        top=Side(style='thin', color='666666'),
        bottom=Side(style='thin', color='666666')
    )
    
    # Headers
    ws.cell(row=1, column=1, value="Generation")
    ws.cell(row=1, column=2, value="Use Case ID")
    
    # Styling headers principali
    for col in [1, 2]:
        cell = ws.cell(row=1, column=col)
        cell.fill = main_header_fill
        cell.font = main_header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thick_border
    
    # Headers test cases TC1-TC10
    for i in range(1, 11):
        col = 2 + i
        cell = ws.cell(row=1, column=col, value=f"TC{i}")
        cell.fill = tc_header_fill
        cell.font = tc_header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thick_border
    
    current_row = 2
    
    # Popola i dati
    for gen_data in organized_data:
        generation = gen_data["generation"]
        use_cases = gen_data["use_cases"]
        
        if not use_cases:
            continue
            
        # Per ogni use case
        for i, uc_data in enumerate(use_cases):
            use_case_id = uc_data["use_case_id"]
            num_test_cases = uc_data["num_test_cases"]
            
            # Generation cell (merge per tutte le righe del generation)
            if i == 0:
                ws.cell(row=current_row, column=1, value=generation)
                if len(use_cases) > 1:
                    ws.merge_cells(start_row=current_row, start_column=1,
                                  end_row=current_row + len(use_cases) - 1, end_column=1)
                
                gen_cell = ws.cell(row=current_row, column=1)
                gen_cell.fill = PatternFill(
                    start_color=generation_colors.get(generation, "F0F0F0"),
                    end_color=generation_colors.get(generation, "F0F0F0"),
                    fill_type="solid"
                )
                gen_cell.font = Font(bold=True, size=11)
                gen_cell.alignment = Alignment(horizontal='center', vertical='center')
                gen_cell.border = thick_border
            
            # Use Case cell
            uc_cell = ws.cell(row=current_row, column=2, value=use_case_id)
            uc_cell.fill = use_case_fill
            uc_cell.font = use_case_font
            uc_cell.alignment = Alignment(horizontal='center', vertical='center')
            uc_cell.border = thick_border
            
            # Test case cells TC1-TC10
            for tc_num in range(1, 11):
                col = 2 + tc_num
                tc_cell = ws.cell(row=current_row, column=col, value="")
                
                # Se questo test case è presente (entro il numero di test case di questo UC)
                if tc_num <= num_test_cases:
                    tc_cell.fill = present_tc_fill
                    tc_cell.border = thick_border
                else:
                    tc_cell.border = thin_border
                
                tc_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Dropdown Yes/No per tutti i test case
                try:
                    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
                    dv.showDropDown = True
                    dv.showErrorMessage = True
                    ws.add_data_validation(dv)
                    dv.add(tc_cell)
                except Exception as e:
                    print(f"Errore validazione: {e}")
                    pass
            
            current_row += 1
    
    # Ottimizzazione colonne con larghezza standard
    ws.column_dimensions['A'].width = 20  # Generation
    ws.column_dimensions['B'].width = 15  # Use Case ID
    
    # Test case columns - larghezza standard
    for tc_num in range(1, 11):
        column_letter = ws.cell(row=1, column=2 + tc_num).column_letter
        ws.column_dimensions[column_letter].width = 8  # Larghezza standard per vedere Yes/No
    
    # Freeze panes
    ws.freeze_panes = 'C2'
    
    # Salva
    output_path = "/Users/raphaelmazzoleni/Desktop/RaphaelMac/ExplorationModule/Con Evaluation/Evaluation/Quantitative/Feasibility/UC-Analysis-Feasibility.xlsx"
    wb.save(output_path)
    
    print(f"File Excel creato: {output_path}")
    print(f"Generazioni processate: {len(organized_data)}")
    print(f"Righe dati: {current_row - 2}")
    
    return output_path

if __name__ == "__main__":
    create_feasibility_excel()

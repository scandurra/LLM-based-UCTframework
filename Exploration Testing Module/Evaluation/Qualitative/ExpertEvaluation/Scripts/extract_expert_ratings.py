"""
Expert Assessment Data Extractor
Extracts expert ratings from Excel files and creates summary sheets
Excludes Expert 3 data and renames Expert 4 as Expert 3
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
from collections import defaultdict
import statistics

# Configuration
UC_MAPPINGS = {
    'UC1': 'UC1-Assessment',
    'UC2.4': 'UC2_4-Assessment',
    'UC3.3': 'UC3_3-Assessment',
    'UC3.4.4': 'UC3_4_4-Assessment'
}

EXPERTS = ['Expert1', 'Expert2', 'Expert4']  # Excluding Expert 3, will rename Expert 4 to Expert 3
EXPERT_DISPLAY_NAMES = {
    'Expert1': 'Expert 1',
    'Expert2': 'Expert 2',
    'Expert4': 'Expert 3'  # Rename Expert 4 to Expert 3
}

QUALITY_DIMENSIONS = [
    'Clarity & Completeness',
    'Semantic consistency',
    'Context Deviation',
    'Level of creativity'
]

STRATEGIES = [
    'R1 Zero Shot', 'R2 Zero Shot', 'R3 Zero Shot',
    'R1 One Shot', 'R2 One Shot', 'R3 One Shot',
    'R1 Few Shot', 'R2 Few Shot', 'R3 Few Shot'
]

# Mapping ratings to numeric values
RATING_MAP = {
    '(1) Poor': 1,
    '(2) Below Average': 2,
    '(3) Average': 3,
    '(4) Good': 4,
    '(5) Excellent': 5
}


def extract_rating_value(cell_value):
    """Extract numeric rating from cell value"""
    if cell_value is None:
        return None
    
    cell_str = str(cell_value).strip()
    
    # Try exact match first
    if cell_str in RATING_MAP:
        return RATING_MAP[cell_str]
    
    # Try to extract rating from string like "(4) Good" or "4"
    for rating_str, rating_val in RATING_MAP.items():
        if rating_str in cell_str:
            return rating_val
    
    # Try to extract just the number
    try:
        num = int(cell_str.split(')')[0].strip('(').strip())
        if 1 <= num <= 5:
            return num
    except:
        pass
    
    return None


def load_assessments_from_file(filepath, uc_name):
    """Load assessment data from a single Excel file"""
    wb = openpyxl.load_workbook(filepath)
    data = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
    
    for expert_sheet in EXPERTS:
        ws = wb[expert_sheet]
        
        # Read ratings for each dimension and strategy
        for dim_idx, dimension in enumerate(QUALITY_DIMENSIONS):
            row_idx = 5 + dim_idx  # Dimensions start at row 5
            
            for strat_idx, strategy in enumerate(STRATEGIES):
                col_idx = 2 + strat_idx  # Strategies start at column B (2)
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                rating = extract_rating_value(cell_value)
                
                expert_display = EXPERT_DISPLAY_NAMES[expert_sheet]
                data[expert_display][uc_name][dimension][strategy] = rating
    
    wb.close()
    return data


def load_all_assessments():
    """Load all assessment data from all UC files"""
    base_path = '/Users/raphaelmazzoleni/Desktop/RaphaelMac/ExplorationModule/Con Evaluation/Evaluation/Qualitative/ExpertEvaluation'
    
    all_data = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
    
    for uc_name, uc_dir in UC_MAPPINGS.items():
        filepath = os.path.join(base_path, uc_dir, f'{uc_dir.split("-")[0]}-Assessment.xlsx')
        print(f"Loading {filepath}...")
        
        file_data = load_assessments_from_file(filepath, uc_name)
        
        # Merge data
        for expert, uc_data in file_data.items():
            for uc, dim_data in uc_data.items():
                for dim, strat_data in dim_data.items():
                    all_data[expert][uc][dim].update(strat_data)
    
    return all_data


def create_per_expert_summary(all_data, output_dir):
    """Create a summary Excel file for each expert"""
    
    for expert in ['Expert 1', 'Expert 2', 'Expert 3']:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create a sheet for each quality dimension
        for dimension in QUALITY_DIMENSIONS:
            ws = wb.create_sheet(title=dimension[:31])  # Excel sheet name limit
            
            # Headers
            ws['A1'] = f'{expert} - {dimension}'
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:K1')
            
            # Column headers
            ws['A3'] = 'Use Case'
            for col_idx, strategy in enumerate(STRATEGIES, start=2):
                col_letter = get_column_letter(col_idx)
                ws[f'{col_letter}3'] = strategy
                ws[f'{col_letter}3'].font = Font(bold=True)
                ws[f'{col_letter}3'].alignment = Alignment(horizontal='center', wrap_text=True)
            
            ws['K3'] = 'Average'
            ws['K3'].font = Font(bold=True)
            
            # Apply header styling
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            for col in range(1, 12):
                cell = ws.cell(row=3, column=col)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Data rows
            row = 4
            for uc_name in sorted(UC_MAPPINGS.keys()):
                ws.cell(row=row, column=1, value=uc_name)
                ws.cell(row=row, column=1).font = Font(bold=True)
                
                ratings = []
                for col_idx, strategy in enumerate(STRATEGIES, start=2):
                    rating = all_data[expert][uc_name][dimension].get(strategy)
                    ws.cell(row=row, column=col_idx, value=rating)
                    ws.cell(row=row, column=col_idx).alignment = Alignment(horizontal='center')
                    if rating is not None:
                        ratings.append(rating)
                
                # Calculate average
                if ratings:
                    avg = statistics.mean(ratings)
                    ws.cell(row=row, column=11, value=round(avg, 2))
                    ws.cell(row=row, column=11).alignment = Alignment(horizontal='center')
                    ws.cell(row=row, column=11).font = Font(bold=True)
                
                row += 1
            
            # Overall average row
            ws.cell(row=row, column=1, value='Overall Average')
            ws.cell(row=row, column=1).font = Font(bold=True, size=11)
            
            all_ratings = []
            for col_idx, strategy in enumerate(STRATEGIES, start=2):
                strat_ratings = []
                for uc_name in UC_MAPPINGS.keys():
                    rating = all_data[expert][uc_name][dimension].get(strategy)
                    if rating is not None:
                        strat_ratings.append(rating)
                        all_ratings.append(rating)
                
                if strat_ratings:
                    avg = statistics.mean(strat_ratings)
                    ws.cell(row=row, column=col_idx, value=round(avg, 2))
                    ws.cell(row=row, column=col_idx).alignment = Alignment(horizontal='center')
                    ws.cell(row=row, column=col_idx).font = Font(bold=True)
                    ws.cell(row=row, column=col_idx).fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            
            if all_ratings:
                overall_avg = statistics.mean(all_ratings)
                ws.cell(row=row, column=11, value=round(overall_avg, 2))
                ws.cell(row=row, column=11).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=11).font = Font(bold=True, size=11)
                ws.cell(row=row, column=11).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 15
            for col in range(2, 12):
                ws.column_dimensions[get_column_letter(col)].width = 12
        
        # Save workbook
        filename = os.path.join(output_dir, f'{expert.replace(" ", "_")}_Summary.xlsx')
        wb.save(filename)
        print(f"Created: {filename}")


def create_per_use_case_summary(all_data, output_dir):
    """Create a summary Excel file for each use case"""
    
    for uc_name in sorted(UC_MAPPINGS.keys()):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Create a sheet for each quality dimension
        for dimension in QUALITY_DIMENSIONS:
            ws = wb.create_sheet(title=dimension[:31])
            
            # Headers
            ws['A1'] = f'{uc_name} - {dimension}'
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:K1')
            
            # Column headers
            ws['A3'] = 'Expert'
            for col_idx, strategy in enumerate(STRATEGIES, start=2):
                col_letter = get_column_letter(col_idx)
                ws[f'{col_letter}3'] = strategy
                ws[f'{col_letter}3'].font = Font(bold=True)
                ws[f'{col_letter}3'].alignment = Alignment(horizontal='center', wrap_text=True)
            
            ws['K3'] = 'Average'
            ws['K3'].font = Font(bold=True)
            
            # Apply header styling
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            for col in range(1, 12):
                cell = ws.cell(row=3, column=col)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Data rows
            row = 4
            for expert in ['Expert 1', 'Expert 2', 'Expert 3']:
                ws.cell(row=row, column=1, value=expert)
                ws.cell(row=row, column=1).font = Font(bold=True)
                
                ratings = []
                for col_idx, strategy in enumerate(STRATEGIES, start=2):
                    rating = all_data[expert][uc_name][dimension].get(strategy)
                    ws.cell(row=row, column=col_idx, value=rating)
                    ws.cell(row=row, column=col_idx).alignment = Alignment(horizontal='center')
                    if rating is not None:
                        ratings.append(rating)
                
                # Calculate average
                if ratings:
                    avg = statistics.mean(ratings)
                    ws.cell(row=row, column=11, value=round(avg, 2))
                    ws.cell(row=row, column=11).alignment = Alignment(horizontal='center')
                    ws.cell(row=row, column=11).font = Font(bold=True)
                
                row += 1
            
            # Average across experts
            ws.cell(row=row, column=1, value='Average Across Experts')
            ws.cell(row=row, column=1).font = Font(bold=True, size=11)
            
            all_ratings = []
            for col_idx, strategy in enumerate(STRATEGIES, start=2):
                strat_ratings = []
                for expert in ['Expert 1', 'Expert 2', 'Expert 3']:
                    rating = all_data[expert][uc_name][dimension].get(strategy)
                    if rating is not None:
                        strat_ratings.append(rating)
                        all_ratings.append(rating)
                
                if strat_ratings:
                    avg = statistics.mean(strat_ratings)
                    ws.cell(row=row, column=col_idx, value=round(avg, 2))
                    ws.cell(row=row, column=col_idx).alignment = Alignment(horizontal='center')
                    ws.cell(row=row, column=col_idx).font = Font(bold=True)
                    ws.cell(row=row, column=col_idx).fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            
            if all_ratings:
                overall_avg = statistics.mean(all_ratings)
                ws.cell(row=row, column=11, value=round(overall_avg, 2))
                ws.cell(row=row, column=11).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=11).font = Font(bold=True, size=11)
                ws.cell(row=row, column=11).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 20
            for col in range(2, 12):
                ws.column_dimensions[get_column_letter(col)].width = 12
        
        # Save workbook
        filename = os.path.join(output_dir, f'{uc_name.replace(".", "_")}_Summary.xlsx')
        wb.save(filename)
        print(f"Created: {filename}")


def calculate_global_statistics(all_data):
    """Calculate global statistics across all experts, UCs, and dimensions"""
    stats = {
        'by_dimension': {},
        'by_strategy': {},
        'by_expert': {},
        'by_uc': {},
        'overall': {}
    }
    
    # By dimension
    for dimension in QUALITY_DIMENSIONS:
        ratings = []
        for expert in ['Expert 1', 'Expert 2', 'Expert 3']:
            for uc_name in UC_MAPPINGS.keys():
                for strategy in STRATEGIES:
                    rating = all_data[expert][uc_name][dimension].get(strategy)
                    if rating is not None:
                        ratings.append(rating)
        
        if ratings:
            stats['by_dimension'][dimension] = {
                'mean': statistics.mean(ratings),
                'stdev': statistics.stdev(ratings) if len(ratings) > 1 else 0,
                'min': min(ratings),
                'max': max(ratings),
                'count': len(ratings)
            }
    
    # By strategy
    for strategy in STRATEGIES:
        ratings = []
        for expert in ['Expert 1', 'Expert 2', 'Expert 3']:
            for uc_name in UC_MAPPINGS.keys():
                for dimension in QUALITY_DIMENSIONS:
                    rating = all_data[expert][uc_name][dimension].get(strategy)
                    if rating is not None:
                        ratings.append(rating)
        
        if ratings:
            stats['by_strategy'][strategy] = {
                'mean': statistics.mean(ratings),
                'stdev': statistics.stdev(ratings) if len(ratings) > 1 else 0,
                'count': len(ratings)
            }
    
    # By expert
    for expert in ['Expert 1', 'Expert 2', 'Expert 3']:
        ratings = []
        for uc_name in UC_MAPPINGS.keys():
            for dimension in QUALITY_DIMENSIONS:
                for strategy in STRATEGIES:
                    rating = all_data[expert][uc_name][dimension].get(strategy)
                    if rating is not None:
                        ratings.append(rating)
        
        if ratings:
            stats['by_expert'][expert] = {
                'mean': statistics.mean(ratings),
                'stdev': statistics.stdev(ratings) if len(ratings) > 1 else 0,
                'count': len(ratings)
            }
    
    # By use case
    for uc_name in UC_MAPPINGS.keys():
        ratings = []
        for expert in ['Expert 1', 'Expert 2', 'Expert 3']:
            for dimension in QUALITY_DIMENSIONS:
                for strategy in STRATEGIES:
                    rating = all_data[expert][uc_name][dimension].get(strategy)
                    if rating is not None:
                        ratings.append(rating)
        
        if ratings:
            stats['by_uc'][uc_name] = {
                'mean': statistics.mean(ratings),
                'stdev': statistics.stdev(ratings) if len(ratings) > 1 else 0,
                'count': len(ratings)
            }
    
    # Overall
    all_ratings = []
    for expert in ['Expert 1', 'Expert 2', 'Expert 3']:
        for uc_name in UC_MAPPINGS.keys():
            for dimension in QUALITY_DIMENSIONS:
                for strategy in STRATEGIES:
                    rating = all_data[expert][uc_name][dimension].get(strategy)
                    if rating is not None:
                        all_ratings.append(rating)
    
    if all_ratings:
        stats['overall'] = {
            'mean': statistics.mean(all_ratings),
            'stdev': statistics.stdev(all_ratings) if len(all_ratings) > 1 else 0,
            'min': min(all_ratings),
            'max': max(all_ratings),
            'count': len(all_ratings)
        }
    
    return stats


def print_statistics_report(stats):
    """Print a formatted statistics report"""
    print("\n" + "="*80)
    print("EXPERT ASSESSMENT STATISTICS REPORT")
    print("="*80)
    
    print("\n--- OVERALL STATISTICS ---")
    if stats['overall']:
        o = stats['overall']
        print(f"Mean: {o['mean']:.2f}")
        print(f"Std Dev: {o['stdev']:.2f}")
        print(f"Min: {o['min']}, Max: {o['max']}")
        print(f"Total ratings: {o['count']}")
    
    print("\n--- BY QUALITY DIMENSION ---")
    for dim, values in stats['by_dimension'].items():
        print(f"\n{dim}:")
        print(f"  Mean: {values['mean']:.2f} ± {values['stdev']:.2f}")
        print(f"  Range: [{values['min']}, {values['max']}]")
        print(f"  Count: {values['count']}")
    
    print("\n--- BY GENERATION STRATEGY ---")
    for strat, values in stats['by_strategy'].items():
        print(f"\n{strat}:")
        print(f"  Mean: {values['mean']:.2f} ± {values['stdev']:.2f}")
        print(f"  Count: {values['count']}")
    
    print("\n--- BY EXPERT ---")
    for expert, values in stats['by_expert'].items():
        print(f"\n{expert}:")
        print(f"  Mean: {values['mean']:.2f} ± {values['stdev']:.2f}")
        print(f"  Count: {values['count']}")
    
    print("\n--- BY USE CASE ---")
    for uc, values in stats['by_uc'].items():
        print(f"\n{uc}:")
        print(f"  Mean: {values['mean']:.2f} ± {values['stdev']:.2f}")
        print(f"  Count: {values['count']}")
    
    print("\n" + "="*80)


def main():
    """Main execution function"""
    output_dir = '/Users/raphaelmazzoleni/Desktop/RaphaelMac/ExplorationModule/Con Evaluation/Evaluation/Qualitative/ExpertEvaluation/Scripts'
    
    print("Loading all assessment data...")
    all_data = load_all_assessments()
    
    print("\nCreating per-expert summary files...")
    create_per_expert_summary(all_data, output_dir)
    
    print("\nCreating per-use-case summary files...")
    create_per_use_case_summary(all_data, output_dir)
    
    print("\nCalculating statistics...")
    stats = calculate_global_statistics(all_data)
    
    print_statistics_report(stats)
    
    print("\n✓ All files created successfully!")
    print(f"Output directory: {output_dir}")


if __name__ == "__main__":
    main()

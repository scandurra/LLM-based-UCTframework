import os
import json
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import matplotlib.pyplot as plt
import seaborn as sns
from collections import defaultdict

class JSONSyntaxAnalyzer:
    def __init__(self, base_path):
        self.base_path = Path(base_path)
        self.generations_path = self.base_path / "Generations"
        self.quantitative_path = self.base_path / "Evaluation" / "Quantitative"
        self.json_syntax_path = self.quantitative_path / "JSON Syntax"
        self.charts_path = self.json_syntax_path / "Charts"
        
        # Create directories
        self.json_syntax_path.mkdir(parents=True, exist_ok=True)
        self.charts_path.mkdir(parents=True, exist_ok=True)
        
        # Required JSON fields
        self.required_fields = [
            "test_case_id", "title", "preconditions", "postconditions", 
            "test_steps", "test_type", "priority", "use_case_id"
        ]
        
        self.required_step_fields = ["step", "expected"]
        
        # Generation types
        self.generation_types = [
            "R1 Zero Shot", "R1 One Shot", "R1 Few Shot",
            "R2 Zero Shot", "R2 One Shot", "R2 Few Shot",
            "R3 Zero Shot", "R3 One Shot", "R3 Few Shot"
        ]
        
        self.prompt_types = ["Zero Shot", "One Shot", "Few Shot"]
        
    def analyze_file_formats(self):
        """Analyze file format distribution across generations"""
        
        print("📊 ANALYZING FILE FORMATS...")
        
        format_analysis = {}
        
        for gen_type in self.generation_types:
            gen_folder = self.generations_path / gen_type
            
            if not gen_folder.exists():
                continue
            
            json_count = len(list(gen_folder.glob("*.json")))
            txt_count = len(list(gen_folder.glob("*.txt")))
            total_count = json_count + txt_count
            
            json_percentage = (json_count / total_count * 100) if total_count > 0 else 0
            txt_percentage = (txt_count / total_count * 100) if total_count > 0 else 0
            
            format_analysis[gen_type] = {
                'json_count': json_count,
                'txt_count': txt_count,
                'total_count': total_count,
                'json_percentage': json_percentage,
                'txt_percentage': txt_percentage
            }
            
            print(f"  {gen_type}: JSON={json_count}, TXT={txt_count}, Total={total_count}")
        
        return format_analysis
    
    def analyze_prompt_types(self, format_analysis):
        """Analyze format distribution by prompt type"""
        
        print("📊 ANALYZING BY PROMPT TYPE...")
        
        prompt_analysis = {}
        
        for prompt_type in self.prompt_types:
            json_total = 0
            txt_total = 0
            
            for gen_type, data in format_analysis.items():
                if prompt_type in gen_type:
                    json_total += data['json_count']
                    txt_total += data['txt_count']
            
            total = json_total + txt_total
            json_percentage = (json_total / total * 100) if total > 0 else 0
            txt_percentage = (txt_total / total * 100) if total > 0 else 0
            
            prompt_analysis[prompt_type] = {
                'json_count': json_total,
                'txt_count': txt_total,
                'total_count': total,
                'json_percentage': json_percentage,
                'txt_percentage': txt_percentage
            }
            
            print(f"  {prompt_type}: JSON={json_total}, TXT={txt_total}, Total={total}")
        
        return prompt_analysis
    
    def validate_json_structure(self, json_data):
        """Validate JSON structure against required fields"""
        
        errors = []
        
        if not isinstance(json_data, list):
            errors.append("Root should be an array")
            return errors
        
        for i, test_case in enumerate(json_data):
            if not isinstance(test_case, dict):
                errors.append(f"Test case {i+1} should be an object")
                continue
            
            # Check required fields
            for field in self.required_fields:
                if field not in test_case:
                    errors.append(f"Test case {i+1}: Missing field '{field}'")
                elif test_case[field] is None or test_case[field] == "":
                    errors.append(f"Test case {i+1}: Empty field '{field}'")
            
            # Check test_steps structure
            if "test_steps" in test_case:
                if not isinstance(test_case["test_steps"], list):
                    errors.append(f"Test case {i+1}: 'test_steps' should be an array")
                else:
                    for j, step in enumerate(test_case["test_steps"]):
                        if not isinstance(step, dict):
                            errors.append(f"Test case {i+1}, step {j+1}: Should be an object")
                            continue
                        
                        for step_field in self.required_step_fields:
                            if step_field not in step:
                                errors.append(f"Test case {i+1}, step {j+1}: Missing field '{step_field}'")
                            elif step[step_field] is None or step[step_field] == "":
                                errors.append(f"Test case {i+1}, step {j+1}: Empty field '{step_field}'")
        
        return errors
    
    def analyze_json_content(self):
        """Analyze JSON content for structural errors"""
        
        print("🔍 ANALYZING JSON CONTENT STRUCTURE...")
        
        content_analysis = {}
        
        for gen_type in self.generation_types:
            gen_folder = self.generations_path / gen_type
            
            if not gen_folder.exists():
                continue
            
            gen_errors = defaultdict(list)
            total_files = 0
            valid_files = 0
            syntax_error_files = 0
            structure_error_files = 0
            
            for file_path in gen_folder.glob("*.json"):
                total_files += 1
                
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        json_data = json.load(f)
                    
                    # Check structure
                    errors = self.validate_json_structure(json_data)
                    
                    if errors:
                        structure_error_files += 1
                        gen_errors[file_path.name] = errors
                    else:
                        valid_files += 1
                        
                except json.JSONDecodeError as e:
                    syntax_error_files += 1
                    gen_errors[file_path.name] = [f"JSON Syntax Error: {str(e)}"]
                except Exception as e:
                    syntax_error_files += 1
                    gen_errors[file_path.name] = [f"File Error: {str(e)}"]
            
            # Also check .txt files (these are JSON syntax errors)
            for file_path in gen_folder.glob("*.txt"):
                total_files += 1
                syntax_error_files += 1
                gen_errors[file_path.name] = ["JSON Syntax Error: Saved as TXT due to invalid JSON"]
            
            content_analysis[gen_type] = {
                'total_files': total_files,
                'valid_files': valid_files,
                'syntax_error_files': syntax_error_files,
                'structure_error_files': structure_error_files,
                'errors': dict(gen_errors),
                'valid_percentage': (valid_files / total_files * 100) if total_files > 0 else 0,
                'syntax_error_percentage': (syntax_error_files / total_files * 100) if total_files > 0 else 0,
                'structure_error_percentage': (structure_error_files / total_files * 100) if total_files > 0 else 0
            }
            
            print(f"  {gen_type}: Valid={valid_files}, Syntax Errors={syntax_error_files}, Structure Errors={structure_error_files}")
        
        return content_analysis
    
    def create_excel_report(self, format_analysis, prompt_analysis, content_analysis):
        """Create comprehensive Excel report"""
        
        print("📄 CREATING EXCEL REPORT...")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "JSON Syntax Analysis"
        
        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        section_font = Font(bold=True, size=11)
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        current_row = 1
        
        # Title
        ws.merge_cells(f'A{current_row}:H{current_row}')
        title_cell = ws.cell(row=current_row, column=1, value="JSON Syntax Analysis Report")
        title_cell.fill = header_fill
        title_cell.font = header_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        current_row += 2
        
        # Section 1: File Format Analysis by Generation
        ws.merge_cells(f'A{current_row}:H{current_row}')
        section_cell = ws.cell(row=current_row, column=1, value="File Format Analysis by Generation")
        section_cell.fill = section_fill
        section_cell.font = section_font
        
        current_row += 1
        
        # Headers
        headers = ['Generation Type', 'JSON Files', 'TXT Files', 'Total Files', 'JSON %', 'TXT %', 'Syntax Error Rate']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = thin_border
        
        current_row += 1
        
        # Data
        for gen_type, data in format_analysis.items():
            row_data = [
                gen_type, data['json_count'], data['txt_count'], data['total_count'],
                f"{data['json_percentage']:.1f}%", f"{data['txt_percentage']:.1f}%",
                f"{data['txt_percentage']:.1f}%"
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = thin_border
            
            current_row += 1
        
        current_row += 2
        
        # Section 2: Analysis by Prompt Type
        ws.merge_cells(f'A{current_row}:H{current_row}')
        section_cell = ws.cell(row=current_row, column=1, value="Analysis by Prompt Type")
        section_cell.fill = section_fill
        section_cell.font = section_font
        
        current_row += 1
        
        # Headers
        for col, header in enumerate(headers[:-1], 1):  # Same headers except last one
            cell = ws.cell(row=current_row, column=col, value=header.replace('Generation Type', 'Prompt Type'))
            cell.font = Font(bold=True)
            cell.border = thin_border
        
        current_row += 1
        
        # Data
        for prompt_type, data in prompt_analysis.items():
            row_data = [
                prompt_type, data['json_count'], data['txt_count'], data['total_count'],
                f"{data['json_percentage']:.1f}%", f"{data['txt_percentage']:.1f}%"
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = thin_border
            
            current_row += 1
        
        current_row += 2
        
        # Section 3: JSON Structure Analysis
        ws.merge_cells(f'A{current_row}:I{current_row}')
        section_cell = ws.cell(row=current_row, column=1, value="JSON Structure Analysis")
        section_cell.fill = section_fill
        section_cell.font = section_font
        
        current_row += 1
        
        # Headers
        structure_headers = ['Generation Type', 'Total Files', 'Valid Files', 'Syntax Errors', 
                           'Structure Errors', 'Valid %', 'Syntax Error %', 'Structure Error %']
        for col, header in enumerate(structure_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = thin_border
        
        current_row += 1
        
        # Data
        for gen_type, data in content_analysis.items():
            row_data = [
                gen_type, data['total_files'], data['valid_files'], 
                data['syntax_error_files'], data['structure_error_files'],
                f"{data['valid_percentage']:.1f}%", 
                f"{data['syntax_error_percentage']:.1f}%",
                f"{data['structure_error_percentage']:.1f}%"
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = thin_border
            
            current_row += 1
        
        # Adjust column widths
        column_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        for i, letter in enumerate(column_letters):
            ws.column_dimensions[letter].width = 18
        
        # Create second sheet with detailed errors
        error_ws = wb.create_sheet(title="Detailed Errors")
        
        error_row = 1
        error_ws.cell(row=error_row, column=1, value="Generation Type").font = Font(bold=True)
        error_ws.cell(row=error_row, column=2, value="File Name").font = Font(bold=True)
        error_ws.cell(row=error_row, column=3, value="Error Description").font = Font(bold=True)
        
        error_row += 1
        
        for gen_type, data in content_analysis.items():
            for file_name, errors in data['errors'].items():
                for error in errors:
                    error_ws.cell(row=error_row, column=1, value=gen_type)
                    error_ws.cell(row=error_row, column=2, value=file_name)
                    error_ws.cell(row=error_row, column=3, value=error)
                    error_row += 1
        
        # Adjust error sheet columns
        error_ws.column_dimensions['A'].width = 20
        error_ws.column_dimensions['B'].width = 25
        error_ws.column_dimensions['C'].width = 60
        
        # Save file
        excel_path = self.json_syntax_path / "UC-JSON-Analysis-Summary.xlsx"
        wb.save(excel_path)
        print(f"✅ Excel report saved: {excel_path}")
        
        return excel_path
    
    def create_charts(self, format_analysis, prompt_analysis, content_analysis):
        """Create visualization charts"""
        
        print("📊 CREATING VISUALIZATION CHARTS...")
        
        # Set style
        plt.style.use('default')
        sns.set_palette("husl")
        
        # Chart 1: File Format Distribution by Generation
        fig, ax = plt.subplots(figsize=(14, 8))
        
        generations = list(format_analysis.keys())
        json_counts = [format_analysis[gen]['json_count'] for gen in generations]
        txt_counts = [format_analysis[gen]['txt_count'] for gen in generations]
        
        x = np.arange(len(generations))
        width = 0.35
        
        bars1 = ax.bar(x - width/2, json_counts, width, label='JSON (Valid Syntax)', color='lightgreen')
        bars2 = ax.bar(x + width/2, txt_counts, width, label='TXT (Syntax Errors)', color='lightcoral')
        
        ax.set_title('File Format Distribution by Generation Type', fontsize=16, fontweight='bold')
        ax.set_xlabel('Generation Type', fontsize=12)
        ax.set_ylabel('Number of Files', fontsize=12)
        ax.set_xticks(x)
        ax.set_xticklabels(generations, rotation=45, ha='right')
        ax.legend()
        ax.grid(True, alpha=0.3)
        
        # Add value labels
        for bars in [bars1, bars2]:
            for bar in bars:
                height = bar.get_height()
                if height > 0:
                    ax.text(bar.get_x() + bar.get_width()/2., height,
                           f'{int(height)}', ha='center', va='bottom')
        
        plt.tight_layout()
        plt.savefig(self.charts_path / '1_file_format_by_generation.png', dpi=300, bbox_inches='tight')
        plt.close()
        
        # Chart 2: Syntax Error Rate by Prompt Type
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 6))
        
        # Pie chart
        prompt_types = list(prompt_analysis.keys())
        error_counts = [prompt_analysis[pt]['txt_count'] for pt in prompt_types]
        colors = ['lightblue', 'lightgreen', 'lightcoral']
        
        wedges, texts, autotexts = ax1.pie(error_counts, labels=prompt_types, autopct='%1.1f%%',
                                          colors=colors, startangle=90)
        ax1.set_title('Syntax Error Distribution by Prompt Type', fontweight='bold')
        
        # Bar chart with percentages
        json_percentages = [prompt_analysis[pt]['json_percentage'] for pt in prompt_types]
        txt_percentages = [prompt_analysis[pt]['txt_percentage'] for pt in prompt_types]
        
        x = np.arange(len(prompt_types))
        bars1 = ax2.bar(x, json_percentages, width, label='Valid JSON %', color='lightgreen')
        bars2 = ax2.bar(x, txt_percentages, width, bottom=json_percentages, 
                       label='Syntax Error %', color='lightcoral')
        
        ax2.set_title('Success Rate by Prompt Type', fontweight='bold')
        ax2.set_xlabel('Prompt Type')
        ax2.set_ylabel('Percentage')
        ax2.set_xticks(x)
        ax2.set_xticklabels(prompt_types)
        ax2.legend()
        ax2.set_ylim(0, 100)
        
        # Add percentage labels
        for i, (json_pct, txt_pct) in enumerate(zip(json_percentages, txt_percentages)):
            if json_pct > 5:
                ax2.text(i, json_pct/2, f'{json_pct:.1f}%', ha='center', va='center', fontweight='bold')
            if txt_pct > 5:
                ax2.text(i, json_pct + txt_pct/2, f'{txt_pct:.1f}%', ha='center', va='center', fontweight='bold')
        
        plt.tight_layout()
        plt.savefig(self.charts_path / '2_syntax_error_analysis.png', dpi=300, bbox_inches='tight')
        plt.close()
        
        # Chart 3: Comprehensive Quality Analysis
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(16, 12))
        
        # Structure error analysis
        generations = list(content_analysis.keys())
        valid_percentages = [content_analysis[gen]['valid_percentage'] for gen in generations]
        syntax_error_percentages = [content_analysis[gen]['syntax_error_percentage'] for gen in generations]
        structure_error_percentages = [content_analysis[gen]['structure_error_percentage'] for gen in generations]
        
        x = np.arange(len(generations))
        
        ax1.bar(x, valid_percentages, width, label='Valid', color='lightgreen')
        ax1.bar(x, syntax_error_percentages, width, bottom=valid_percentages, 
               label='Syntax Errors', color='lightcoral')
        ax1.bar(x, structure_error_percentages, width, 
               bottom=np.array(valid_percentages) + np.array(syntax_error_percentages),
               label='Structure Errors', color='orange')
        
        ax1.set_title('JSON Quality Analysis by Generation', fontweight='bold')
        ax1.set_xlabel('Generation Type')
        ax1.set_ylabel('Percentage')
        ax1.set_xticks(x)
        ax1.set_xticklabels(generations, rotation=45, ha='right')
        ax1.legend()
        ax1.set_ylim(0, 100)
        
        # Error type distribution
        total_syntax_errors = sum(content_analysis[gen]['syntax_error_files'] for gen in generations)
        total_structure_errors = sum(content_analysis[gen]['structure_error_files'] for gen in generations)
        
        error_types = ['Syntax Errors', 'Structure Errors']
        error_counts = [total_syntax_errors, total_structure_errors]
        
        ax2.pie(error_counts, labels=error_types, autopct='%1.1f%%', 
               colors=['lightcoral', 'orange'], startangle=90)
        ax2.set_title('Error Type Distribution', fontweight='bold')
        
        # Trend analysis by round
        rounds = ['R1', 'R2', 'R3']
        round_error_rates = []
        
        for round_num in rounds:
            round_gens = [gen for gen in generations if gen.startswith(round_num)]
            if round_gens:
                total_files = sum(content_analysis[gen]['total_files'] for gen in round_gens)
                total_errors = sum(content_analysis[gen]['syntax_error_files'] + 
                                 content_analysis[gen]['structure_error_files'] for gen in round_gens)
                error_rate = (total_errors / total_files * 100) if total_files > 0 else 0
                round_error_rates.append(error_rate)
            else:
                round_error_rates.append(0)
        
        ax3.plot(rounds, round_error_rates, marker='o', linewidth=2, markersize=8, color='red')
        ax3.set_title('Error Rate Trend by Round', fontweight='bold')
        ax3.set_xlabel('Generation Round')
        ax3.set_ylabel('Error Rate (%)')
        ax3.grid(True, alpha=0.3)
        ax3.set_ylim(0, max(round_error_rates) * 1.1 if round_error_rates else 1)
        
        # Add value labels
        for i, rate in enumerate(round_error_rates):
            ax3.text(i, rate, f'{rate:.1f}%', ha='center', va='bottom')
        
        # Success rate comparison
        prompt_success_rates = [prompt_analysis[pt]['json_percentage'] for pt in prompt_types]
        
        ax4.bar(prompt_types, prompt_success_rates, color=['lightblue', 'lightgreen', 'lightcoral'])
        ax4.set_title('JSON Generation Success Rate', fontweight='bold')
        ax4.set_xlabel('Prompt Type')
        ax4.set_ylabel('Success Rate (%)')
        ax4.set_ylim(0, 100)
        
        # Add value labels
        for i, rate in enumerate(prompt_success_rates):
            ax4.text(i, rate, f'{rate:.1f}%', ha='center', va='bottom')
        
        plt.tight_layout()
        plt.savefig(self.charts_path / '3_comprehensive_quality_analysis.png', dpi=300, bbox_inches='tight')
        plt.close()
        
        print(f"✅ Charts saved to: {self.charts_path}")
    
    def run_analysis(self):
        """Run complete JSON syntax analysis"""
        
        print("🚀 STARTING JSON SYNTAX ANALYSIS")
        print("=" * 50)
        
        # Step 1: Analyze file formats
        format_analysis = self.analyze_file_formats()
        
        # Step 2: Analyze by prompt type
        prompt_analysis = self.analyze_prompt_types(format_analysis)
        
        # Step 3: Analyze JSON content structure
        content_analysis = self.analyze_json_content()
        
        # Step 4: Create Excel report
        excel_path = self.create_excel_report(format_analysis, prompt_analysis, content_analysis)
        
        # Step 5: Create charts
        self.create_charts(format_analysis, prompt_analysis, content_analysis)
        
        # Print summary
        print("\n📊 ANALYSIS SUMMARY")
        print("=" * 50)
        
        total_files = sum(data['total_count'] for data in format_analysis.values())
        total_json = sum(data['json_count'] for data in format_analysis.values())
        total_txt = sum(data['txt_count'] for data in format_analysis.values())
        
        print(f"Total files analyzed: {total_files}")
        print(f"Valid JSON files: {total_json} ({total_json/total_files*100:.1f}%)")
        print(f"Syntax error files (TXT): {total_txt} ({total_txt/total_files*100:.1f}%)")
        
        print(f"\nBy Prompt Type:")
        for prompt_type, data in prompt_analysis.items():
            print(f"  {prompt_type}: {data['json_percentage']:.1f}% success rate")
        
        print(f"\nFiles created:")
        print(f"  📄 {excel_path}")
        print(f"  📊 {len(list(self.charts_path.glob('*.png')))} charts in {self.charts_path}")

if __name__ == "__main__":
    base_path = "c:/Users/user/Desktop/Progetti/ExplorationModule/Con Evaluation"
    analyzer = JSONSyntaxAnalyzer(base_path)
    analyzer.run_analysis()

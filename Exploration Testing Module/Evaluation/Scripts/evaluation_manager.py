import os
import json
import openpyxl
from shutil import copy2
import re
from datetime import datetime
import time


class EvaluationManager:
    """
    Manages the evaluation of generated test cases.
    Updates Excel analysis files for each Use Case.
    """
    
    def __init__(self, base_path="."):
        self.base_path = base_path
        self.template_path = os.path.join(base_path, "UCX-Analysis-Template.xlsx")
        self.evaluation_folder = os.path.join(base_path, "Evaluation")
        self.generations_output = os.path.join(base_path, "Generations", "Output")
        self.generations_log = os.path.join(base_path, "Generations", "Log")
        self.final_output = os.path.join(base_path, "FinalOutput")
        
        # Mapping for rounds and prompting types
        self.round_types = [
            "R1 Zero Shot", "R2 Zero Shot", "R3 Zero Shot",
            "R1 One Shot", "R2 One Shot", "R3 One Shot", 
            "R1 Few Shot", "R2 Few Shot", "R3 Few Shot"
        ]
        
    def get_uc_evaluation_folder(self, uc_id):
        """Gets the evaluation folder path for a specific UC"""
        uc_folder = f"{uc_id}-Evaluation"
        return os.path.join(self.evaluation_folder, uc_folder)
    
    def get_uc_excel_path(self, uc_id):
        """Gets the analysis Excel file path for a specific UC"""
        uc_folder = self.get_uc_evaluation_folder(uc_id)
        return os.path.join(uc_folder, f"{uc_id}-Analysis.xlsx")
    
    def initialize_excel_files(self, force=False):
        """
        Initializes all analysis Excel files by copying the template.
        
        Args:
            force (bool): If True, overwrites existing files. 
                         If False, creates only missing files.
        """
        print("🔄 Initializing analysis Excel files...")
        
        if not os.path.exists(self.template_path):
            print(f"❌ Template not found: {self.template_path}")
            return False
        
        # Get all UCs from evaluation folders
        uc_folders = [f for f in os.listdir(self.evaluation_folder) 
                     if f.endswith("-Evaluation") and os.path.isdir(os.path.join(self.evaluation_folder, f))]
        
        created_count = 0
        skipped_count = 0
        
        for uc_folder in uc_folders:
            uc_id = uc_folder.replace("-Evaluation", "")
            excel_path = self.get_uc_excel_path(uc_id)
            
            # Create folder if it doesn't exist
            os.makedirs(os.path.dirname(excel_path), exist_ok=True)
            
            # Check if file already exists
            if os.path.exists(excel_path) and not force:
                skipped_count += 1
                continue
            
            # Copy template
            copy2(self.template_path, excel_path)
            
            # Update file with UC ID
            self._update_uc_id_in_excel(excel_path, uc_id)
            
            print(f"✅ {'Recreated' if force else 'Initialized'}: {excel_path}")
            created_count += 1
        
        if force:
            print(f"🎯 Recreated {created_count} analysis Excel files")
        else:
            print(f"🎯 Created {created_count} new Excel files, {skipped_count} existing kept")
        
        return True
    
    def _update_uc_id_in_excel(self, excel_path, uc_id):
        """Updates the UC ID in the Excel file"""
        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            
            # Find UCX cell and replace it with real ID
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == "UCX":
                        cell.value = uc_id
                        break
            
            wb.save(excel_path)
        except Exception as e:
            print(f"⚠️ Error updating UC ID in {excel_path}: {e}")
    
    def detect_round_and_type(self, current_generation_folder=None):
        """
        Detects the current round and prompting type based on the most recent log files
        or the current generation folder if specified.
        """
        if current_generation_folder:
            # Extract information from folder name
            folder_name = os.path.basename(current_generation_folder)
            for round_type in self.round_types:
                if round_type.replace(" ", "_").lower() in folder_name.lower():
                    return round_type
        
        # Fallback: search in most recent logs
        if not os.path.exists(self.generations_log):
            return "R1 Zero Shot"  # Default
        
        log_files = [f for f in os.listdir(self.generations_log) if f.startswith("log_") and f.endswith(".txt")]
        if not log_files:
            return "R1 Zero Shot"  # Default
        
        # Sort by modification date (most recent first)
        log_files.sort(key=lambda x: os.path.getmtime(os.path.join(self.generations_log, x)), reverse=True)
        
        # Analyze log file names to determine type
        for log_file in log_files[:10]:  # Check the 10 most recent
            if "ZS" in log_file:
                if "R1" in log_file:
                    return "R1 Zero Shot"
                elif "R2" in log_file:
                    return "R2 Zero Shot"
                elif "R3" in log_file:
                    return "R3 Zero Shot"
            elif "1S" in log_file:
                if "R1" in log_file:
                    return "R1 One Shot"
                elif "R2" in log_file:
                    return "R2 One Shot"
                elif "R3" in log_file:
                    return "R3 One Shot"
            elif "FS" in log_file:
                if "R1" in log_file:
                    return "R1 Few Shot"
                elif "R2" in log_file:
                    return "R2 Few Shot"
                elif "R3" in log_file:
                    return "R3 Few Shot"
        
        return "R1 Zero Shot"  # Default
    
    def count_test_cases(self, uc_id, round_type=None):
        """Counts the number of test cases generated for a UC"""
        # If we have a specific round type, search first in that folder
        if round_type:
            round_folder = self._convert_round_type_to_folder_name(round_type)
            generations_dir = os.path.join(self.base_path, "Generations", round_folder)
            
            if os.path.exists(generations_dir):
                # First search for JSON files
                json_file = os.path.join(generations_dir, f"{uc_id}.json")
                if os.path.exists(json_file):
                    return self._count_from_file(json_file)
                
                # If no JSON, search for TXT files
                txt_file = os.path.join(generations_dir, f"{uc_id}.txt")
                if os.path.exists(txt_file):
                    return self._count_from_file(txt_file)
        
        # Fallback: First try in FinalOutput
        final_output_path = os.path.join(self.final_output, f"{uc_id}.json")
        
        if os.path.exists(final_output_path):
            return self._count_from_file(final_output_path)
        
        # Then try in general generations output
        gen_output_path = os.path.join(self.generations_output, f"{uc_id}.json")
        if os.path.exists(gen_output_path):
            return self._count_from_file(gen_output_path)
        
        # Finally search in specific round folders
        generations_dir = os.path.join(self.base_path, "Generations")
        if os.path.exists(generations_dir):
            for item in os.listdir(generations_dir):
                item_path = os.path.join(generations_dir, item)
                if os.path.isdir(item_path) and "Shot" in item:
                    # First search for JSON files
                    json_file = os.path.join(item_path, f"{uc_id}.json")
                    if os.path.exists(json_file):
                        return self._count_from_file(json_file)
                    
                    # If no JSON, search for TXT files
                    txt_file = os.path.join(item_path, f"{uc_id}.txt")
                    if os.path.exists(txt_file):
                        return self._count_from_file(txt_file)
        
        return 0
    
    def _count_from_file(self, file_path):
        """Helper to count test cases from a JSON file"""
        try:
            # First try to parse as valid JSON
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read().strip()
                
            try:
                data = json.loads(content)
                if isinstance(data, list):
                    return len(data)
                else:
                    return 1 if data else 0
            except json.JSONDecodeError:
                # If JSON is not valid, manually count test_case_id
                test_case_count = content.count('"test_case_id":')
                if test_case_count > 0:
                    return test_case_count
                
                # Fallback: count UCX_TCY patterns
                import re
                pattern = r'"test_case_id":\s*"UC\d+_TC\d+"'
                matches = re.findall(pattern, content)
                return len(matches)
                
        except Exception as e:
            print(f"⚠️ Error counting test cases from {file_path}: {e}")
            return 0
    
    def check_need_repair(self, uc_id, round_type=None):
        """Checks if the file needed repair for a specific round"""
        # If we have a specific round type, check only that folder
        if round_type:
            round_folder = self._convert_round_type_to_folder_name(round_type)
            generations_dir = os.path.join(self.base_path, "Generations", round_folder)
            
            if os.path.exists(generations_dir):
                json_file = os.path.join(generations_dir, f"{uc_id}.json")
                txt_file = os.path.join(generations_dir, f"{uc_id}.txt")
                
                # If JSON file exists, no repair needed
                if os.path.exists(json_file):
                    return "no"
                # If only TXT file exists, repair needed
                elif os.path.exists(txt_file):
                    return "yes"
        
        # Fallback: First check in general output folder
        gen_output_txt = os.path.join(self.generations_output, f"{uc_id}.txt")
        if os.path.exists(gen_output_txt):
            return "yes"
        
        # Then check in specific round folders
        generations_dir = os.path.join(self.base_path, "Generations")
        if os.path.exists(generations_dir):
            for item in os.listdir(generations_dir):
                item_path = os.path.join(generations_dir, item)
                if os.path.isdir(item_path) and "Shot" in item:
                    txt_file = os.path.join(item_path, f"{uc_id}.txt")
                    if os.path.exists(txt_file):
                        return "yes"
        
        return "no"
    
    def get_execution_time(self, uc_id, round_type):
        """Gets execution time from corresponding log file"""
        # Convert round type to log format
        log_suffix = self._convert_round_type_to_log_format(round_type)
        
        # Search for most recent log file for this UC and round type
        if not os.path.exists(self.generations_log):
            return 0.0
        
        log_files = [f for f in os.listdir(self.generations_log) 
                    if f.startswith(f"log_{uc_id}_{log_suffix}") and f.endswith(".txt")]
        
        if not log_files:
            return 0.0
        
        # Take the most recent file
        log_files.sort(key=lambda x: os.path.getmtime(os.path.join(self.generations_log, x)), reverse=True)
        log_path = os.path.join(self.generations_log, log_files[0])
        
        try:
            with open(log_path, 'r', encoding='utf-8') as f:
                content = f.read()
                # Search for line with EXECUTION TIME
                match = re.search(r'⏱️ EXECUTION TIME: ([\d.]+) seconds', content)
                if match:
                    return float(match.group(1))
        except Exception as e:
            print(f"⚠️ Error reading execution time from {log_path}: {e}")
        
        return 0.0
    
    def get_token_counts(self, uc_id, round_type):
        """
        Gets prompt and response token counts from corresponding log file.
        Returns a tuple (prompt_tokens, response_tokens).
        """
        # Convert round type to log format
        log_suffix = self._convert_round_type_to_log_format(round_type)
        
        # Search for most recent log file for this UC and round type
        if not os.path.exists(self.generations_log):
            return 0, 0
        
        log_files = [f for f in os.listdir(self.generations_log) 
                    if f.startswith(f"log_{uc_id}_{log_suffix}") and f.endswith(".txt")]
        
        if not log_files:
            return 0, 0
        
        # Take the most recent file
        log_files.sort(key=lambda x: os.path.getmtime(os.path.join(self.generations_log, x)), reverse=True)
        log_path = os.path.join(self.generations_log, log_files[0])
        
        try:
            with open(log_path, 'r', encoding='utf-8') as f:
                content = f.read()
                # Search for line with TOKENS USED
                match = re.search(r'🪙 TOKENS USED: (.+)', content)
                if match:
                    token_info_str = match.group(1)
                    # If it's a dictionary, try to extract prompt_tokens and completion_tokens
                    if 'prompt_tokens' in token_info_str and 'completion_tokens' in token_info_str:
                        prompt_match = re.search(r"'prompt_tokens': (\d+)", token_info_str)
                        completion_match = re.search(r"'completion_tokens': (\d+)", token_info_str)
                        if prompt_match and completion_match:
                            return int(prompt_match.group(1)), int(completion_match.group(1))
                    # Fallback: if it's just a number, consider it as total_tokens
                    try:
                        total_tokens = int(token_info_str)
                        # If we don't have separation, return 0, 0
                        return 0, 0
                    except:
                        pass
        except Exception as e:
            print(f"⚠️ Error reading token count from {log_path}: {e}")
        
        return 0, 0

    def get_token_count(self, uc_id, round_type):
        """
        Deprecated method. Use get_token_counts() to get prompt and response tokens separately.
        Maintained for compatibility, returns the sum of prompt + response tokens.
        """
        prompt_tokens, response_tokens = self.get_token_counts(uc_id, round_type)
        return prompt_tokens + response_tokens
    
    def get_number_of_examples(self, round_type):
        """Gets the number of examples based on prompting type"""
        if "Zero Shot" in round_type:
            return 0
        elif "One Shot" in round_type:
            return 1
        elif "Few Shot" in round_type:
            return 2  # Few Shot always uses 2 examples
        return 0
    
    def _convert_round_type_to_folder_name(self, round_type):
        """Converts round type to generation folder name"""
        return round_type  # "R1 Zero Shot" -> "R1 Zero Shot"
    
    def _convert_round_type_to_log_format(self, round_type):
        """Converts round type to format used in log file names"""
        mapping = {
            "R1 Zero Shot": "ZS_R1",
            "R2 Zero Shot": "ZS_R2", 
            "R3 Zero Shot": "ZS_R3",
            "R1 One Shot": "1S_R1",
            "R2 One Shot": "1S_R2",
            "R3 One Shot": "1S_R3",
            "R1 Few Shot": "FS_R1",
            "R2 Few Shot": "FS_R2",
            "R3 Few Shot": "FS_R3"
        }
        return mapping.get(round_type, "ZS_R1")
    
    def update_evaluation_for_uc(self, uc_id, round_type):
        """Updates evaluation for a specific UC and round type"""
        excel_path = self.get_uc_excel_path(uc_id)
        
        if not os.path.exists(excel_path):
            print(f"❌ Excel file not found: {excel_path}")
            return False
        
        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            
            # Find row corresponding to round type
            target_row = None
            for row_num in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row_num, column=1).value
                if cell_value == round_type:
                    target_row = row_num
                    break
            
            if target_row is None:
                print(f"⚠️ Round type '{round_type}' not found in Excel file")
                return False
            
            # Collect data
            n_test_cases = self.count_test_cases(uc_id, round_type)
            n_examples = self.get_number_of_examples(round_type)
            need_repair = self.check_need_repair(uc_id, round_type)
            execution_time = self.get_execution_time(uc_id, round_type)
            prompt_tokens, response_tokens = self.get_token_counts(uc_id, round_type)
            
            # Update cells (columns 2-7 based on new template structure)
            ws.cell(row=target_row, column=2).value = n_test_cases    # N. Test cases
            ws.cell(row=target_row, column=3).value = n_examples      # N. Examples
            ws.cell(row=target_row, column=4).value = need_repair     # Need to repair
            ws.cell(row=target_row, column=5).value = execution_time  # Execution time
            ws.cell(row=target_row, column=6).value = prompt_tokens   # Token prompt
            ws.cell(row=target_row, column=7).value = response_tokens # Token response
            
            wb.save(excel_path)
            
            print(f"✅ Updated {uc_id} - {round_type}:")
            print(f"   📊 Test cases: {n_test_cases}")
            print(f"   📝 Examples: {n_examples}")
            print(f"   🔧 Need repair: {need_repair}")
            print(f"   ⏱️ Execution time: {execution_time:.2f}s")
            print(f"   🪙 Token prompt: {prompt_tokens}")
            print(f"   🪙 Token response: {response_tokens}")
            print(f"   🪙 Total tokens: {prompt_tokens + response_tokens}")
            
            return True
            
        except Exception as e:
            print(f"❌ Error updating {excel_path}: {e}")
            return False
    
    def update_all_evaluations(self, round_type=None):
        """Updates all evaluations for current round type"""
        if round_type is None:
            round_type = self.detect_round_and_type()
        
        print(f"🔄 Updating evaluations for: {round_type}")
        
        # Get all UCs from evaluation folders
        uc_folders = [f for f in os.listdir(self.evaluation_folder) 
                     if f.endswith("-Evaluation") and os.path.isdir(os.path.join(self.evaluation_folder, f))]
        
        updated_count = 0
        for uc_folder in uc_folders:
            uc_id = uc_folder.replace("-Evaluation", "")
            if self.update_evaluation_for_uc(uc_id, round_type):
                updated_count += 1
        
        print(f"🎯 Updated {updated_count}/{len(uc_folders)} evaluations")
        return updated_count
    
    def log_evaluation_update(self, uc_id, round_type, data):
        """Logs what was written to files"""
        log_message = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] EVALUATION UPDATE - {uc_id} - {round_type}"
        log_message += f"\n  📊 Test cases: {data.get('n_test_cases', 0)}"
        log_message += f"\n  📝 Examples: {data.get('n_examples', 0)}"
        log_message += f"\n  🔧 Need repair: {data.get('need_repair', 'no')}"
        log_message += f"\n  ⏱️ Execution time: {data.get('execution_time', 0):.2f}s"
        log_message += f"\n  🪙 Token prompt: {data.get('prompt_tokens', 0)}"
        log_message += f"\n  🪙 Token response: {data.get('response_tokens', 0)}"
        log_message += f"\n  🪙 Total tokens: {data.get('prompt_tokens', 0) + data.get('response_tokens', 0)}"
        log_message += f"\n  📄 Excel file: {self.get_uc_excel_path(uc_id)}"
        
        print(log_message)
        
        # Optionally save to dedicated log file
        eval_log_path = os.path.join(self.evaluation_folder, "evaluation_log.txt")
        with open(eval_log_path, 'a', encoding='utf-8') as f:
            f.write(log_message + "\n\n")


def main():
    """Main function for testing and direct usage"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Evaluation Manager for Test Case Generation')
    parser.add_argument('--init', action='store_true', help='Initialize all analysis Excel files')
    parser.add_argument('--update', type=str, help='Update evaluations for a specific round type')
    parser.add_argument('--update-all', action='store_true', help='Update all evaluations for current round')
    parser.add_argument('--uc', type=str, help='Specific UC to update')
    
    args = parser.parse_args()
    
    eval_manager = EvaluationManager()
    
    if args.init:
        eval_manager.initialize_excel_files()
    elif args.update:
        if args.uc:
            eval_manager.update_evaluation_for_uc(args.uc, args.update)
        else:
            eval_manager.update_all_evaluations(args.update)
    elif args.update_all:
        eval_manager.update_all_evaluations()
    else:
        print("Usage: python evaluation_manager.py [--init | --update ROUND_TYPE | --update-all] [--uc UC_ID]")
        print("Examples:")
        print("  python evaluation_manager.py --init")
        print("  python evaluation_manager.py --update 'R1 Zero Shot'")
        print("  python evaluation_manager.py --update-all")
        print("  python evaluation_manager.py --update 'R2 Few Shot' --uc UC1")


if __name__ == "__main__":
    main()

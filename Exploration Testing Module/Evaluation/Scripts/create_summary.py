import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from pathlib import Path
import openpyxl

def create_summary_excel():
    """Create a proper UC-Analysis-Summary.xlsx file"""
    
    base_path = Path("c:/Users/user/Desktop/Progetti/ExplorationModule/Con Evaluation")
    quantitative_path = base_path / "Evaluation" / "Quantitative"
    summary_path = quantitative_path / "Summary"
    template_path = base_path / "Evaluation" / "UCX-Analysis-Template.xlsx"
    
    # Collect all data
    all_data = []
    xlsx_files = list(quantitative_path.glob("*.xlsx"))
    xlsx_files = [f for f in xlsx_files if not f.name.startswith("~$") and "Summary" not in str(f)]
    
    print(f"Processing {len(xlsx_files)} files for summary...")
    
    for file_path in xlsx_files:
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            use_case = file_path.stem
            
            # Read data starting from row 4
            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                if i >= 4 and row[0] and str(row[0]) != 'nan':
                    generation_type = str(row[0])
                    n_test_cases = int(row[1]) if row[1] and str(row[1]) != 'nan' and str(row[1]).isdigit() else 0
                    n_examples = int(row[2]) if row[2] and str(row[2]) != 'nan' and str(row[2]).isdigit() else 0
                    need_repair = str(row[3]) if row[3] and str(row[3]) != 'nan' else 'unknown'
                    
                    # Handle interference latency conversion
                    latency_val = row[4]
                    if latency_val and str(latency_val) != 'nan':
                        try:
                            interference_latency = float(latency_val)
                        except:
                            interference_latency = 0
                    else:
                        interference_latency = 0
                    
                    token_prompt = int(row[5]) if row[5] and str(row[5]) != 'nan' and str(row[5]).isdigit() else 0
                    token_response = int(row[6]) if row[6] and str(row[6]) != 'nan' and str(row[6]).isdigit() else 0
                    
                    # Extract shot type and round
                    if 'Zero Shot' in generation_type:
                        shot_type = 'Zero Shot'
                        round_num = generation_type.split()[0]
                    elif 'One Shot' in generation_type:
                        shot_type = 'One Shot'
                        round_num = generation_type.split()[0]
                    elif 'Few Shot' in generation_type:
                        shot_type = 'Few Shot'
                        round_num = generation_type.split()[0]
                    else:
                        continue
                    
                    all_data.append({
                        'use_case': use_case,
                        'generation_type': generation_type,
                        'round': round_num,
                        'shot_type': shot_type,
                        'n_test_cases': n_test_cases,
                        'n_examples': n_examples,
                        'need_repair': need_repair,
                        'interference_latency': interference_latency,
                        'token_prompt': token_prompt,
                        'token_response': token_response
                    })
                    
        except Exception as e:
            print(f"Error processing {file_path}: {e}")
    
    # Convert to DataFrame
    df = pd.DataFrame(all_data)
    
    if df.empty:
        print("No data collected!")
        return
    
    print(f"Collected {len(df)} data points")
    
    # Create summary statistics by generation type
    summary_stats = df.groupby('generation_type').agg({
        'n_test_cases': ['mean', 'std', 'min', 'max'],
        'interference_latency': ['mean', 'std', 'min', 'max'],
        'token_prompt': ['mean', 'std'],
        'token_response': ['mean', 'std'],
        'need_repair': lambda x: (x == 'yes').sum()
    }).round(2)
    
    # Flatten column names
    summary_stats.columns = ['_'.join(col).strip() for col in summary_stats.columns]
    summary_stats = summary_stats.reset_index()
    
    # Create new workbook based on template
    try:
        template_wb = load_workbook(template_path)
        summary_wb = Workbook()
        
        # Copy template structure
        template_ws = template_wb.active
        summary_ws = summary_wb.active
        summary_ws.title = "Summary Analysis"
        
        # Copy some basic formatting from template
        for row in range(1, 6):  # Copy first few rows for structure
            for col in range(1, 8):
                template_cell = template_ws.cell(row=row, column=col)
                summary_cell = summary_ws.cell(row=row, column=col)
                
                if template_cell.value:
                    summary_cell.value = template_cell.value
                
                # Copy basic formatting
                if template_cell.has_style:
                    summary_cell.font = Font(
                        name=template_cell.font.name,
                        size=template_cell.font.size,
                        bold=template_cell.font.bold
                    )
        
        # Override with summary content
        summary_ws['A1'] = "Generation Analysis Summary"
        summary_ws['A2'] = "Aggregated Results Across All Use Cases"
        
        # Add summary statistics starting from row 4
        start_row = 4
        
        # Headers
        headers = ['Generation Type', 'Avg Test Cases', 'Avg Latency [s]', 'Std Latency', 
                  'Min Latency', 'Max Latency', 'Avg Prompt Tokens', 'Avg Response Tokens', 'Repairs Needed']
        
        for col, header in enumerate(headers, 1):
            cell = summary_ws.cell(row=start_row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        for row_idx, (_, row_data) in enumerate(summary_stats.iterrows(), 1):
            summary_ws.cell(row=start_row + row_idx, column=1, value=row_data['generation_type'])
            summary_ws.cell(row=start_row + row_idx, column=2, value=row_data['n_test_cases_mean'])
            summary_ws.cell(row=start_row + row_idx, column=3, value=row_data['interference_latency_mean'])
            summary_ws.cell(row=start_row + row_idx, column=4, value=row_data['interference_latency_std'])
            summary_ws.cell(row=start_row + row_idx, column=5, value=row_data['interference_latency_min'])
            summary_ws.cell(row=start_row + row_idx, column=6, value=row_data['interference_latency_max'])
            summary_ws.cell(row=start_row + row_idx, column=7, value=row_data['token_prompt_mean'])
            summary_ws.cell(row=start_row + row_idx, column=8, value=row_data['token_response_mean'])
            summary_ws.cell(row=start_row + row_idx, column=9, value=row_data['need_repair_<lambda>'])
        
        # Add overall summary
        overall_start = start_row + len(summary_stats) + 3
        summary_ws.cell(row=overall_start, column=1, value="OVERALL SUMMARY").font = Font(bold=True, size=14)
        
        overall_stats = [
            ("Total Use Cases Analyzed", len(df['use_case'].unique())),
            ("Total Data Points", len(df)),
            ("Average Interference Latency", f"{df['interference_latency'].mean():.2f} ± {df['interference_latency'].std():.2f} seconds"),
            ("Best Performing (Lowest Avg Latency)", summary_stats.loc[summary_stats['interference_latency_mean'].idxmin(), 'generation_type']),
            ("Most Consistent (Lowest Std)", summary_stats.loc[summary_stats['interference_latency_std'].idxmin(), 'generation_type']),
            ("Highest Test Case Generation", summary_stats.loc[summary_stats['n_test_cases_mean'].idxmax(), 'generation_type'])
        ]
        
        for i, (label, value) in enumerate(overall_stats):
            summary_ws.cell(row=overall_start + i + 1, column=1, value=label)
            summary_ws.cell(row=overall_start + i + 1, column=2, value=str(value))
        
        # Adjust column widths
        for col in range(1, 10):
            summary_ws.column_dimensions[summary_ws.cell(row=1, column=col).column_letter].width = 15
        
        # Save the file
        summary_file_path = summary_path / "UC-Analysis-Summary.xlsx"
        summary_wb.save(summary_file_path)
        print(f"Created comprehensive summary: {summary_file_path}")
        
        # Also create detailed breakdown by shot type and round
        detailed_summary = df.groupby(['round', 'shot_type']).agg({
            'n_test_cases': 'mean',
            'interference_latency': ['mean', 'std'],
            'token_prompt': 'mean',
            'token_response': 'mean',
            'need_repair': lambda x: (x == 'yes').sum()
        }).round(2)
        
        # Add detailed breakdown sheet
        detail_ws = summary_wb.create_sheet(title="Detailed Breakdown")
        detail_ws['A1'] = "Detailed Analysis by Round and Shot Type"
        detail_ws['A1'].font = Font(bold=True, size=14)
        
        # Convert detailed summary to a more readable format
        detailed_data = []
        for (round_num, shot_type), data in detailed_summary.iterrows():
            detailed_data.append([
                round_num, shot_type,
                data[('n_test_cases', 'mean')],
                data[('interference_latency', 'mean')],
                data[('interference_latency', 'std')],
                data[('token_prompt', 'mean')],
                data[('token_response', 'mean')],
                data[('need_repair', '<lambda>')]
            ])
        
        # Add headers for detailed breakdown
        detail_headers = ['Round', 'Shot Type', 'Avg Test Cases', 'Avg Latency [s]', 
                         'Std Latency', 'Avg Prompt Tokens', 'Avg Response Tokens', 'Repairs Needed']
        
        for col, header in enumerate(detail_headers, 1):
            cell = detail_ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add detailed data
        for row_idx, row_data in enumerate(detailed_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                detail_ws.cell(row=3 + row_idx, column=col_idx, value=value)
        
        summary_wb.save(summary_file_path)
        print(f"Added detailed breakdown sheet")
        
    except Exception as e:
        print(f"Error creating summary Excel: {e}")
        
        # Fallback: create simple Excel file
        simple_summary = pd.DataFrame({
            'Generation_Type': summary_stats['generation_type'],
            'Avg_Test_Cases': summary_stats['n_test_cases_mean'],
            'Avg_Interference_Latency': summary_stats['interference_latency_mean'],
            'Std_Interference_Latency': summary_stats['interference_latency_std'],
            'Avg_Prompt_Tokens': summary_stats['token_prompt_mean'],
            'Avg_Response_Tokens': summary_stats['token_response_mean'],
            'Repairs_Needed': summary_stats['need_repair_<lambda>']
        })
        
        simple_path = summary_path / "UC-Analysis-Summary-Simple.xlsx"
        simple_summary.to_excel(simple_path, index=False)
        print(f"Created fallback summary: {simple_path}")

if __name__ == "__main__":
    create_summary_excel()

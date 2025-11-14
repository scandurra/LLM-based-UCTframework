import os
import json
from typing import List, Dict, Tuple
from pathlib import Path
import numpy as np
import pandas as pd
from sklearn.metrics.pairwise import cosine_similarity
from sentence_transformers import SentenceTransformer
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Carica il modello locale per gli embeddings
MODEL_PATH = "paraphrase-multilingual-mpnet-base-v2"
model = SentenceTransformer(MODEL_PATH)

def load_json_file(filepath: str) -> Dict:
    """Load a JSON or TXT file (containing JSON)"""
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
            
            # Try to fix common JSON issues in txt files
            # Remove trailing commas before closing brackets
            content = content.replace(',]', ']').replace(',}', '}')
            
            # Try to parse
            data = json.loads(content)
            
            # If it's a list, take the first element (the test case)
            if isinstance(data, list) and len(data) > 0:
                return data[0]
            return data
    except json.JSONDecodeError as e:
        # If JSON is incomplete, try to find and extract just the first complete test case
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # Try to find the first complete test case object
            start = content.find('{')
            if start == -1:
                print(f"Error loading {filepath}: No JSON object found")
                return None
            
            # Find matching closing brace
            brace_count = 0
            in_string = False
            escape_next = False
            
            for i in range(start, len(content)):
                char = content[i]
                
                if escape_next:
                    escape_next = False
                    continue
                    
                if char == '\\':
                    escape_next = True
                    continue
                    
                if char == '"' and not escape_next:
                    in_string = not in_string
                    continue
                    
                if not in_string:
                    if char == '{':
                        brace_count += 1
                    elif char == '}':
                        brace_count -= 1
                        if brace_count == 0:
                            # Found complete object
                            json_str = content[start:i+1]
                            return json.loads(json_str)
            
            print(f"Error loading {filepath}: Could not find complete JSON object")
            return None
            
        except Exception as e2:
            print(f"Error loading {filepath}: {e2}")
            return None
    except Exception as e:
        print(f"Error loading {filepath}: {e}")
        return None

def test_case_to_text(test_case: Dict) -> str:
    """Convert a test case to text for embedding"""
    if not test_case:
        return ""
    
    text_parts = []
    
    # Add all relevant fields
    if 'title' in test_case:
        text_parts.append(f"Title: {test_case['title']}")
    
    if 'preconditions' in test_case:
        text_parts.append(f"Preconditions: {test_case['preconditions']}")
    
    if 'postconditions' in test_case:
        text_parts.append(f"Postconditions: {test_case['postconditions']}")
    
    if 'test_steps' in test_case:
        steps_text = "Test steps: "
        for i, step in enumerate(test_case['test_steps'], 1):
            step_desc = step.get('step', '')
            expected = step.get('expected', '')
            steps_text += f"{i}. {step_desc} -> {expected}; "
        text_parts.append(steps_text)
    
    if 'test_type' in test_case:
        text_parts.append(f"Type: {test_case['test_type']}")
    
    if 'priority' in test_case:
        text_parts.append(f"Priority: {test_case['priority']}")
    
    return " ".join(text_parts)

def generate_embedding(text: str) -> np.ndarray:
    """Generate embedding for text using local model"""
    return model.encode(text, convert_to_numpy=True)

def calculate_cosine_sim(emb1: np.ndarray, emb2: np.ndarray) -> float:
    """Calculate cosine similarity between two embeddings"""
    return cosine_similarity([emb1], [emb2])[0][0]

def load_all_test_cases() -> Dict[str, Dict[str, Tuple[Dict, np.ndarray]]]:
    """
    Load all test cases and generate embeddings
    Returns: {
        'baseline': {'UC1_TC2': (test_case_dict, embedding), ...},
        'R1 Few Shot': {'UC1_TC2': (test_case_dict, embedding), ...},
        ...
    }
    """
    base_path = Path(r"c:\Users\user\Desktop\CosineSimilarityEmbedding")
    
    all_data = {}
    
    # Load baseline
    print("Loading baseline...")
    baseline_dir = base_path / "Baseline JSON"
    baseline_data = {}
    
    for file in baseline_dir.glob("*.json"):
        test_case = load_json_file(str(file))
        if test_case:
            tc_id = file.stem  # UC1_TC2, UC24_TC1, etc.
            text = test_case_to_text(test_case)
            embedding = generate_embedding(text)
            baseline_data[tc_id] = (test_case, embedding)
    
    all_data['Baseline'] = baseline_data
    print(f"  Loaded {len(baseline_data)} baseline test cases")
    
    # Load all strategies and rounds
    generated_dir = base_path / "Generated"
    
    for strategy_dir in generated_dir.iterdir():
        if strategy_dir.is_dir():
            strategy_name = strategy_dir.name
            print(f"\nLoading {strategy_name}...")
            strategy_data = {}
            
            # Load both .json and .txt
            for file in list(strategy_dir.glob("*.json")) + list(strategy_dir.glob("*.txt")):
                test_case = load_json_file(str(file))
                if test_case:
                    tc_id = file.stem
                    text = test_case_to_text(test_case)
                    embedding = generate_embedding(text)
                    strategy_data[tc_id] = (test_case, embedding)
            
            all_data[strategy_name] = strategy_data
            print(f"  Loaded {len(strategy_data)} test cases")
    
    return all_data

def calculate_all_similarities(all_data: Dict) -> pd.DataFrame:
    """
    Calculate all cosine similarities between baseline and each strategy
    """
    baseline_data = all_data['Baseline']
    
    results = []
    
    for strategy_name, strategy_data in all_data.items():
        if strategy_name == 'Baseline':
            continue
        
        # Extract Round and Shot from strategy (e.g., "R1 Few Shot" -> Round=1, Shot="Few Shot")
        parts = strategy_name.split()
        round_num = parts[0] if parts else ""
        shot_type = " ".join(parts[1:]) if len(parts) > 1 else ""
        
        for tc_id in baseline_data.keys():
            if tc_id in strategy_data:
                baseline_emb = baseline_data[tc_id][1]
                strategy_emb = strategy_data[tc_id][1]
                
                similarity = calculate_cosine_sim(baseline_emb, strategy_emb)
                
                results.append({
                    'Test Case': tc_id,
                    'Round': round_num,
                    'Strategy': shot_type,
                    'Cosine Similarity': similarity
                })
            else:
                print(f"  WARNING: {tc_id} not found in {strategy_name}")
    
    return pd.DataFrame(results)

def create_summary_tables(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Create summary tables by strategy and round
    """
    # Average by strategy
    summary_by_strategy = df.groupby('Strategy')['Cosine Similarity'].agg([
        ('Mean', 'mean'),
        ('Min', 'min'),
        ('Max', 'max'),
        ('Std Dev', 'std'),
        ('Count', 'count')
    ]).round(4)
    
    # Average by round
    summary_by_round = df.groupby('Round')['Cosine Similarity'].agg([
        ('Mean', 'mean'),
        ('Min', 'min'),
        ('Max', 'max'),
        ('Std Dev', 'std'),
        ('Count', 'count')
    ]).round(4)
    
    # Average by Round + Strategy combination
    summary_by_both = df.groupby(['Round', 'Strategy'])['Cosine Similarity'].agg([
        ('Mean', 'mean'),
        ('Min', 'min'),
        ('Max', 'max'),
        ('Std Dev', 'std'),
        ('Count', 'count')
    ]).round(4)
    
    return summary_by_strategy, summary_by_round, summary_by_both

def format_excel_file(filepath: str, title: str):
    """Format Excel file with styles for better readability"""
    wb = load_workbook(filepath)
    ws = wb.active
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    
    index_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    index_font = Font(name='Calibri', size=10, bold=True)
    
    cell_font = Font(name='Calibri', size=10)
    alignment_center = Alignment(horizontal='center', vertical='center')
    alignment_left = Alignment(horizontal='left', vertical='center')
    
    border_thin = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Apply header formatting
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment_center
        cell.border = border_thin
    
    # Apply index (first column) formatting
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        cell.fill = index_fill
        cell.font = index_font
        cell.alignment = alignment_left
        cell.border = border_thin
    
    # Apply data cell formatting
    for row in range(2, ws.max_row + 1):
        for col in range(2, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = cell_font
            cell.alignment = alignment_center
            cell.border = border_thin
            
            # Format numbers to 4 decimal places
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.0000'
    
    # Auto-adjust column widths
    for col in range(1, ws.max_column + 1):
        max_length = 0
        column = get_column_letter(col)
        
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Freeze first row
    ws.freeze_panes = 'A2'
    
    # Set sheet title
    ws.title = title[:31]  # Excel sheet name limit is 31 characters
    
    wb.save(filepath)

def save_dataframe_to_excel(df: pd.DataFrame, filepath: str, title: str, include_index: bool = True):
    """Save DataFrame to Excel with formatting"""
    # Save to Excel
    df.to_excel(filepath, index=include_index, engine='openpyxl')
    
    # Apply formatting
    format_excel_file(filepath, title)

def main():
    """Main function"""
    print("="*80)
    print("COSINE SIMILARITY ANALYSIS - BASELINE vs GENERATED TEST CASES")
    print("="*80)
    
    # Load all data
    print("\n1. DATA LOADING AND EMBEDDINGS GENERATION")
    print("-"*80)
    all_data = load_all_test_cases()
    
    # Calculate all similarities
    print("\n2. COSINE SIMILARITY CALCULATION")
    print("-"*80)
    df_results = calculate_all_similarities(all_data)
    
    # Show complete table
    print("\n3. COMPLETE RESULTS TABLE")
    print("-"*80)
    print(df_results.to_string(index=False))
    
    # Create pivot tables for better visualization
    print("\n4. PIVOT TABLE: Test Case x (Round + Strategy)")
    print("-"*80)
    df_pivot = df_results.pivot_table(
        values='Cosine Similarity',
        index='Test Case',
        columns=['Round', 'Strategy'],
        aggfunc='mean'
    ).round(4)
    print(df_pivot.to_string())
    
    # Create summary tables
    print("\n5. SUMMARY TABLES")
    print("="*80)
    
    summary_strategy, summary_round, summary_both = create_summary_tables(df_results)
    
    print("\n5.1 SUMMARY BY STRATEGY (Shot Type)")
    print("-"*80)
    print(summary_strategy.to_string())
    
    print("\n5.2 SUMMARY BY ROUND")
    print("-"*80)
    print(summary_round.to_string())
    
    print("\n5.3 SUMMARY BY ROUND x STRATEGY")
    print("-"*80)
    print(summary_both.to_string())
    
    # Save results to CSV and Excel in Results folder
    print("\n6. SAVING RESULTS")
    print("-"*80)
    
    base_dir = Path(r"c:\Users\user\Desktop\CosineSimilarityEmbedding")
    output_dir = base_dir / "Results"
    output_dir.mkdir(exist_ok=True)
    
    # 1. Complete results
    print("Saving complete results...")
    df_results.to_csv(output_dir / "cosine_similarity_results.csv", index=False, encoding='utf-8-sig')
    save_dataframe_to_excel(df_results, str(output_dir / "cosine_similarity_results.xlsx"), 
                            "Complete Results", include_index=False)
    print(f"✓ Complete results saved: CSV + XLSX")
    
    # 2. Pivot table
    print("Saving pivot table...")
    df_pivot.to_csv(output_dir / "cosine_similarity_pivot.csv", encoding='utf-8-sig')
    save_dataframe_to_excel(df_pivot, str(output_dir / "cosine_similarity_pivot.xlsx"), 
                            "Pivot Table", include_index=True)
    print(f"✓ Pivot table saved: CSV + XLSX")
    
    # 3. Summary by strategy
    print("Saving summary by strategy...")
    summary_strategy.to_csv(output_dir / "summary_by_strategy.csv", encoding='utf-8-sig')
    save_dataframe_to_excel(summary_strategy, str(output_dir / "summary_by_strategy.xlsx"), 
                            "Summary by Strategy", include_index=True)
    print(f"✓ Summary by strategy saved: CSV + XLSX")
    
    # 4. Summary by round
    print("Saving summary by round...")
    summary_round.to_csv(output_dir / "summary_by_round.csv", encoding='utf-8-sig')
    save_dataframe_to_excel(summary_round, str(output_dir / "summary_by_round.xlsx"), 
                            "Summary by Round", include_index=True)
    print(f"✓ Summary by round saved: CSV + XLSX")
    
    # 5. Summary by round x strategy
    print("Saving summary by round x strategy...")
    summary_both.to_csv(output_dir / "summary_by_round_strategy.csv", encoding='utf-8-sig')
    save_dataframe_to_excel(summary_both, str(output_dir / "summary_by_round_strategy.xlsx"), 
                            "Summary Round x Strategy", include_index=True)
    print(f"✓ Summary by round x strategy saved: CSV + XLSX")
    
    print("\n" + "-"*80)
    print(f"📊 Total files created: 10 (5 CSV + 5 XLSX)")
    print(f"📁 Location: {output_dir}")
    
    print("\n" + "="*80)
    print("ANALYSIS COMPLETED!")
    print("="*80)

if __name__ == "__main__":
    main()
    
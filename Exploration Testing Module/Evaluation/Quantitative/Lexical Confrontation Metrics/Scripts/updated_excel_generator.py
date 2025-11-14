import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import json

class UpdatedExcelGenerator:
    def __init__(self, base_path):
        self.base_path = Path(base_path)
        self.results_path = self.base_path / "Dataset" / "Results"
        
        # Load the main results
        self.main_results_path = self.results_path / "Metrics_Comparison_Results.xlsx"
        self.df = pd.read_excel(self.main_results_path, sheet_name='All Results')
        
        # Styles
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=12)
        self.section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.section_font = Font(bold=True, size=11)
        self.thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    def create_comprehensive_summary(self):
        """Create comprehensive summary Excel file"""
        
        print("📊 Creating Comprehensive Summary...")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Executive Summary"
        
        current_row = 1
        
        # Title
        ws.merge_cells(f'A{current_row}:H{current_row}')
        title_cell = ws.cell(row=current_row, column=1, value="Comprehensive Metrics Analysis Summary")
        title_cell.fill = self.header_fill
        title_cell.font = self.header_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        current_row += 2
        
        # Overall Statistics
        ws.merge_cells(f'A{current_row}:H{current_row}')
        section_cell = ws.cell(row=current_row, column=1, value="Overall Performance Statistics")
        section_cell.fill = self.section_fill
        section_cell.font = self.section_font
        
        current_row += 1
        
        # Statistics table
        stats_data = [
            ['Metric', 'Mean', 'Standard Deviation', 'Minimum', 'Maximum', 'Count'],
            ['BLEU Score', f"{self.df['bleu'].mean():.6f}", f"{self.df['bleu'].std():.6f}", 
             f"{self.df['bleu'].min():.6f}", f"{self.df['bleu'].max():.6f}", len(self.df)],
            ['ROUGE-1', f"{self.df['rouge1'].mean():.6f}", f"{self.df['rouge1'].std():.6f}", 
             f"{self.df['rouge1'].min():.6f}", f"{self.df['rouge1'].max():.6f}", len(self.df)],
            ['ROUGE-2', f"{self.df['rouge2'].mean():.6f}", f"{self.df['rouge2'].std():.6f}", 
             f"{self.df['rouge2'].min():.6f}", f"{self.df['rouge2'].max():.6f}", len(self.df)],
            ['ROUGE-L', f"{self.df['rougeL'].mean():.6f}", f"{self.df['rougeL'].std():.6f}", 
             f"{self.df['rougeL'].min():.6f}", f"{self.df['rougeL'].max():.6f}", len(self.df)],
            ['Cosine Similarity', f"{self.df['cosine_similarity'].mean():.6f}", f"{self.df['cosine_similarity'].std():.6f}", 
             f"{self.df['cosine_similarity'].min():.6f}", f"{self.df['cosine_similarity'].max():.6f}", len(self.df)]
        ]
        
        for row_data in stats_data:
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                if current_row == len(stats_data) + 2:  # Header row
                    cell.font = Font(bold=True)
                cell.border = self.thin_border
            current_row += 1
        
        current_row += 2
        
        # Performance by Prompt Strategy
        ws.merge_cells(f'A{current_row}:H{current_row}')
        section_cell = ws.cell(row=current_row, column=1, value="Performance by Prompt Strategy")
        section_cell.fill = self.section_fill
        section_cell.font = self.section_font
        
        current_row += 1
        
        prompt_summary = self.df.groupby('prompt_strategy').agg({
            'bleu': ['mean', 'std', 'count'],
            'rouge1': ['mean', 'std'],
            'rouge2': ['mean', 'std'],
            'rougeL': ['mean', 'std'],
            'cosine_similarity': ['mean', 'std']
        }).round(6)
        
        # Headers for prompt analysis
        headers = ['Prompt Strategy', 'Count', 'BLEU Mean', 'BLEU Std', 'ROUGE-1 Mean', 'ROUGE-1 Std',
                  'ROUGE-L Mean', 'ROUGE-L Std', 'Cosine Mean', 'Cosine Std']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = self.thin_border
        current_row += 1
        
        for prompt_strategy in prompt_summary.index:
            row_data = [
                prompt_strategy,
                int(prompt_summary.loc[prompt_strategy, ('bleu', 'count')]),
                f"{prompt_summary.loc[prompt_strategy, ('bleu', 'mean')]:.6f}",
                f"{prompt_summary.loc[prompt_strategy, ('bleu', 'std')]:.6f}",
                f"{prompt_summary.loc[prompt_strategy, ('rouge1', 'mean')]:.6f}",
                f"{prompt_summary.loc[prompt_strategy, ('rouge1', 'std')]:.6f}",
                f"{prompt_summary.loc[prompt_strategy, ('rougeL', 'mean')]:.6f}",
                f"{prompt_summary.loc[prompt_strategy, ('rougeL', 'std')]:.6f}",
                f"{prompt_summary.loc[prompt_strategy, ('cosine_similarity', 'mean')]:.6f}",
                f"{prompt_summary.loc[prompt_strategy, ('cosine_similarity', 'std')]:.6f}"
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = self.thin_border
            current_row += 1
        
        current_row += 2
        
        # Performance by Generation Round
        ws.merge_cells(f'A{current_row}:H{current_row}')
        section_cell = ws.cell(row=current_row, column=1, value="Performance by Generation Round")
        section_cell.fill = self.section_fill
        section_cell.font = self.section_font
        
        current_row += 1
        
        round_summary = self.df.groupby('generation_round').agg({
            'bleu': ['mean', 'std', 'count'],
            'rouge1': ['mean', 'std'],
            'rouge2': ['mean', 'std'],
            'rougeL': ['mean', 'std'],
            'cosine_similarity': ['mean', 'std']
        }).round(6)
        
        # Headers for round analysis
        for col, header in enumerate(headers, 1):
            header_text = header.replace('Prompt Strategy', 'Generation Round')
            cell = ws.cell(row=current_row, column=col, value=header_text)
            cell.font = Font(bold=True)
            cell.border = self.thin_border
        current_row += 1
        
        for gen_round in round_summary.index:
            row_data = [
                gen_round,
                int(round_summary.loc[gen_round, ('bleu', 'count')]),
                f"{round_summary.loc[gen_round, ('bleu', 'mean')]:.6f}",
                f"{round_summary.loc[gen_round, ('bleu', 'std')]:.6f}",
                f"{round_summary.loc[gen_round, ('rouge1', 'mean')]:.6f}",
                f"{round_summary.loc[gen_round, ('rouge1', 'std')]:.6f}",
                f"{round_summary.loc[gen_round, ('rougeL', 'mean')]:.6f}",
                f"{round_summary.loc[gen_round, ('rougeL', 'std')]:.6f}",
                f"{round_summary.loc[gen_round, ('cosine_similarity', 'mean')]:.6f}",
                f"{round_summary.loc[gen_round, ('cosine_similarity', 'std')]:.6f}"
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = self.thin_border
            current_row += 1
        
        # Adjust column widths
        for col in range(1, 11):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        # Save the workbook
        summary_path = self.results_path / "Comprehensive_Analysis_Summary.xlsx"
        wb.save(summary_path)
        print(f"✅ Comprehensive Summary saved: {summary_path}")
        return summary_path
    
    def create_detailed_comparison_matrix(self):
        """Create detailed comparison matrix"""
        
        print("📊 Creating Detailed Comparison Matrix...")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparison Matrix"
        
        current_row = 1
        
        # Title
        ws.merge_cells(f'A{current_row}:J{current_row}')
        title_cell = ws.cell(row=current_row, column=1, value="Detailed Comparison Matrix - All Metrics by Test Case and Generation")
        title_cell.fill = self.header_fill
        title_cell.font = self.header_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        current_row += 2
        
        # Create summary by baseline file
        summary_data = self.df.groupby(['baseline_file', 'generation_type']).agg({
            'bleu': 'mean',
            'rouge1': 'mean',
            'rouge2': 'mean',
            'rougeL': 'mean',
            'cosine_similarity': 'mean'
        }).round(6).reset_index()
        
        # Headers
        headers = ['Baseline File', 'Generation Type', 'BLEU', 'ROUGE-1', 'ROUGE-2', 'ROUGE-L', 'Cosine Similarity']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = self.thin_border
        current_row += 1
        
        # Write data
        for _, row in summary_data.iterrows():
            row_data = [
                row['baseline_file'],
                row['generation_type'],
                f"{row['bleu']:.6f}",
                f"{row['rouge1']:.6f}",
                f"{row['rouge2']:.6f}",
                f"{row['rougeL']:.6f}",
                f"{row['cosine_similarity']:.6f}"
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = self.thin_border
            current_row += 1
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 18
        
        matrix_path = self.results_path / "Detailed_Comparison_Matrix.xlsx"
        wb.save(matrix_path)
        print(f"✅ Comparison Matrix saved: {matrix_path}")
        return matrix_path
    
    def create_statistical_analysis(self):
        """Create statistical analysis workbook"""
        
        print("📊 Creating Statistical Analysis...")
        
        wb = Workbook()
        
        # Sheet 1: Correlation Analysis
        ws1 = wb.active
        ws1.title = "Correlation Analysis"
        
        current_row = 1
        
        # Title
        ws1.merge_cells(f'A{current_row}:F{current_row}')
        title_cell = ws1.cell(row=current_row, column=1, value="Metrics Correlation Analysis")
        title_cell.fill = self.header_fill
        title_cell.font = self.header_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        current_row += 2
        
        # Correlation matrix
        metrics_cols = ['bleu', 'rouge1', 'rouge2', 'rougeL', 'cosine_similarity']
        correlation_matrix = self.df[metrics_cols].corr().round(6)
        
        # Write correlation matrix
        for r in dataframe_to_rows(correlation_matrix, index=True, header=True):
            for col, value in enumerate(r, 1):
                cell = ws1.cell(row=current_row, column=col, value=value)
                if current_row == 3 or col == 1:  # Header row/column
                    cell.font = Font(bold=True)
                cell.border = self.thin_border
            current_row += 1
        
        # Sheet 2: Performance Rankings
        ws2 = wb.create_sheet(title="Performance Rankings")
        
        current_row = 1
        
        # Title
        ws2.merge_cells(f'A{current_row}:G{current_row}')
        title_cell = ws2.cell(row=current_row, column=1, value="Generation Type Performance Rankings")
        title_cell.fill = self.header_fill
        title_cell.font = self.header_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        current_row += 2
        
        # Rankings by metric
        gen_summary = self.df.groupby('generation_type').agg({
            'bleu': 'mean',
            'rouge1': 'mean', 
            'rouge2': 'mean',
            'rougeL': 'mean',
            'cosine_similarity': 'mean'
        }).round(6)
        
        # Create rankings
        rankings = pd.DataFrame()
        for metric in metrics_cols:
            rankings[f'{metric}_rank'] = gen_summary[metric].rank(ascending=False, method='min').astype(int)
        
        # Combine data with rankings
        combined_data = pd.concat([gen_summary, rankings], axis=1)
        
        # Write rankings
        for r in dataframe_to_rows(combined_data, index=True, header=True):
            for col, value in enumerate(r, 1):
                cell = ws2.cell(row=current_row, column=col, value=value)
                if current_row == 3:  # Header row
                    cell.font = Font(bold=True)
                cell.border = self.thin_border
            current_row += 1
        
        # Sheet 3: Best and Worst Performers
        ws3 = wb.create_sheet(title="Best Worst Analysis")
        
        current_row = 1
        
        # Title
        ws3.merge_cells(f'A{current_row}:H{current_row}')
        title_cell = ws3.cell(row=current_row, column=1, value="Best and Worst Performing Combinations")
        title_cell.fill = self.header_fill
        title_cell.font = self.header_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        current_row += 2
        
        # Best performers for each metric
        for metric in metrics_cols:
            ws3.merge_cells(f'A{current_row}:H{current_row}')
            section_cell = ws3.cell(row=current_row, column=1, value=f"Best {metric.upper()} Scores")
            section_cell.fill = self.section_fill
            section_cell.font = self.section_font
            current_row += 1
            
            # Headers
            headers = ['Rank', 'Baseline File', 'Generation Type', 'Score', 'Test Case IDs']
            for col, header in enumerate(headers, 1):
                cell = ws3.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.border = self.thin_border
            current_row += 1
            
            # Top 5 for this metric
            top_5 = self.df.nlargest(5, metric)
            for idx, (_, row) in enumerate(top_5.iterrows(), 1):
                row_data = [
                    idx,
                    row['baseline_file'],
                    row['generation_type'],
                    f"{row[metric]:.6f}",
                    f"{row['baseline_test_id']} vs {row['generated_test_id']}"
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws3.cell(row=current_row, column=col, value=value)
                    cell.border = self.thin_border
                current_row += 1
            
            current_row += 1
        
        # Auto-adjust column widths for all sheets
        for ws in [ws1, ws2, ws3]:
            for col in range(1, 10):
                ws.column_dimensions[chr(64 + col)].width = 15
        
        stats_path = self.results_path / "Statistical_Analysis.xlsx"
        wb.save(stats_path)
        print(f"✅ Statistical Analysis saved: {stats_path}")
        return stats_path
    
    def create_test_case_performance_report(self):
        """Create test case specific performance report"""
        
        print("📊 Creating Test Case Performance Report...")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Case Performance"
        
        current_row = 1
        
        # Title
        ws.merge_cells(f'A{current_row}:I{current_row}')
        title_cell = ws.cell(row=current_row, column=1, value="Individual Test Case Performance Analysis")
        title_cell.fill = self.header_fill
        title_cell.font = self.header_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        current_row += 2
        
        # Group by baseline file and calculate statistics
        baseline_summary = self.df.groupby('baseline_file').agg({
            'bleu': ['mean', 'std', 'min', 'max'],
            'rouge1': ['mean', 'std', 'min', 'max'],
            'rougeL': ['mean', 'std', 'min', 'max'],
            'cosine_similarity': ['mean', 'std', 'min', 'max']
        }).round(6)
        
        # Flatten column names
        baseline_summary.columns = [f'{metric}_{stat}' for metric, stat in baseline_summary.columns]
        
        # Write headers
        headers = ['Baseline File'] + list(baseline_summary.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = self.thin_border
        current_row += 1
        
        # Write data
        for baseline_file in baseline_summary.index:
            row_data = [baseline_file] + [f"{baseline_summary.loc[baseline_file, col]:.6f}" 
                                        for col in baseline_summary.columns]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = self.thin_border
            current_row += 1
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 12
        
        # Add second sheet with detailed breakdown
        ws2 = wb.create_sheet(title="Detailed Breakdown")
        
        # Write the full dataset with better formatting
        for r in dataframe_to_rows(self.df, index=False, header=True):
            for col, value in enumerate(r, 1):
                cell = ws2.cell(row=len(list(ws2.rows)) + 1, column=col, value=value)
                if len(list(ws2.rows)) == 1:  # Header row
                    cell.font = Font(bold=True)
                cell.border = self.thin_border
        
        # Auto-adjust column widths
        for col in range(1, len(self.df.columns) + 1):
            ws2.column_dimensions[chr(64 + col)].width = 15
        
        testcase_path = self.results_path / "Test_Case_Performance_Report.xlsx"
        wb.save(testcase_path)
        print(f"✅ Test Case Performance Report saved: {testcase_path}")
        return testcase_path
    
    def run_complete_excel_generation(self):
        """Generate all Excel files"""
        
        print("🚀 GENERATING UPDATED EXCEL FILES")
        print("=" * 60)
        
        generated_files = []
        
        # 1. Comprehensive Summary
        file1 = self.create_comprehensive_summary()
        generated_files.append(file1)
        
        # 2. Detailed Comparison Matrix
        file2 = self.create_detailed_comparison_matrix()
        generated_files.append(file2)
        
        # 3. Statistical Analysis
        file3 = self.create_statistical_analysis()
        generated_files.append(file3)
        
        # 4. Test Case Performance Report
        file4 = self.create_test_case_performance_report()
        generated_files.append(file4)
        
        print(f"\n✅ EXCEL GENERATION COMPLETE!")
        print("=" * 60)
        print(f"Generated {len(generated_files)} Excel files:")
        for file_path in generated_files:
            print(f"  📄 {file_path.name}")
        
        print(f"\nAll files saved in: {self.results_path}")
        
        return generated_files

if __name__ == "__main__":
    base_path = "c:/Users/user/Desktop/Progetti/ExplorationModule/Con Evaluation/Evaluation/Quantitative/Confrontation metrics"
    
    generator = UpdatedExcelGenerator(base_path)
    generated_files = generator.run_complete_excel_generation()
